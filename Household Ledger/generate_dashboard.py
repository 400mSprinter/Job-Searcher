# Created: 2026-03-15 15:00
"""
가계부 대시보드 생성기
뱅크샐러드 엑셀 데이터를 읽어 인터랙티브 HTML 대시보드를 생성합니다.

사용법: python generate_dashboard.py [엑셀파일경로]
"""
import json
import sys
import os
from datetime import datetime, date, time
from openpyxl import load_workbook

DEFAULT_EXCEL = r'C:\Users\jinih\Downloads\이태리님_2025-03-15~2026-03-15\2025-03-15~2026-03-15.xlsx'


def extract_data(path):
    wb = load_workbook(path, data_only=True)
    ws1 = wb['뱅샐현황']
    ws2 = wb['가계부 내역']

    # 1. Customer info
    customer = {
        'name': ws1['B6'].value or '',
        'gender': ws1['C6'].value or '',
        'age': ws1['D6'].value or 0,
        'creditScore': ws1['E6'].value or 0,
    }

    # 2. Cash flow months
    months = []
    for col in range(5, 18):
        v = ws1.cell(row=11, column=col).value
        if v:
            months.append(str(v))

    # Income categories (rows 12-15)
    income = {}
    for row in range(12, 16):
        cat = ws1.cell(row=row, column=2).value
        if cat and cat != '월수입 총계':
            vals = []
            for col in range(5, 5 + len(months)):
                v = ws1.cell(row=row, column=col).value
                vals.append(v if v else 0)
            income[cat] = vals

    # Expense categories (rows 17-34)
    expense = {}
    for row in range(17, 35):
        cat = ws1.cell(row=row, column=2).value
        if cat and cat not in ('월지출 총계', '순수입 총계'):
            vals = []
            for col in range(5, 5 + len(months)):
                v = ws1.cell(row=row, column=col).value
                vals.append(v if v else 0)
            expense[cat] = vals

    # 3. Assets (rows 43-84: all asset items including real estate & insurance)
    assets = []
    current_cat = ''
    for row in range(43, 85):
        b = ws1.cell(row=row, column=2).value
        c = ws1.cell(row=row, column=3).value
        e = ws1.cell(row=row, column=5).value
        if b and not c:
            current_cat = b
        elif c:
            if b:
                current_cat = b
            amt = e if e else 0
            if isinstance(amt, str):
                amt = 0
            assets.append({
                'id': f'asset-{row}',
                'category': current_cat,
                'name': c,
                'amount': float(amt),
            })

    # 4. Insurance policies
    insurance = []
    for row in range(94, 98):
        company = ws1.cell(row=row, column=2).value
        name = ws1.cell(row=row, column=3).value
        if not name:
            continue
        g = ws1.cell(row=row, column=7).value
        h = ws1.cell(row=row, column=8).value
        insurance.append({
            'company': company or '',
            'name': name,
            'status': ws1.cell(row=row, column=5).value or '',
            'totalPaid': ws1.cell(row=row, column=6).value or 0,
            'startDate': g.strftime('%Y-%m-%d') if hasattr(g, 'strftime') else '',
            'endDate': h.strftime('%Y-%m-%d') if hasattr(h, 'strftime') else '',
        })

    # 5. Investments
    investments = []
    for row in range(104, 113):
        name = ws1.cell(row=row, column=4).value
        if not name:
            continue
        i_val = ws1.cell(row=row, column=9).value
        j_val = ws1.cell(row=row, column=10).value
        investments.append({
            'type': ws1.cell(row=row, column=2).value or '',
            'company': ws1.cell(row=row, column=3).value or '',
            'name': name,
            'principal': float(ws1.cell(row=row, column=6).value or 0),
            'currentValue': float(ws1.cell(row=row, column=7).value or 0),
            'returnRate': float(ws1.cell(row=row, column=8).value or 0),
            'startDate': i_val.strftime('%Y-%m-%d') if hasattr(i_val, 'strftime') else '',
            'endDate': j_val.strftime('%Y-%m-%d') if hasattr(j_val, 'strftime') else '',
        })

    # 6. Transactions
    transactions = []
    for row in range(2, ws2.max_row + 1):
        d = ws2.cell(row=row, column=1).value
        if not d:
            continue
        t = ws2.cell(row=row, column=2).value
        transactions.append({
            'date': d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)[:10],
            'time': str(t) if t else '',
            'type': ws2.cell(row=row, column=3).value or '',
            'category': ws2.cell(row=row, column=4).value or '',
            'subcategory': ws2.cell(row=row, column=5).value or '',
            'description': ws2.cell(row=row, column=6).value or '',
            'amount': float(ws2.cell(row=row, column=7).value or 0),
            'currency': ws2.cell(row=row, column=8).value or 'KRW',
            'paymentMethod': ws2.cell(row=row, column=9).value or '',
            'memo': ws2.cell(row=row, column=10).value or '',
        })

    return {
        'customer': customer,
        'months': months,
        'income': income,
        'expense': expense,
        'assets': assets,
        'insurance': insurance,
        'investments': investments,
        'transactions': transactions,
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }


def generate_html(all_data):
    """all_data is a list of person dicts (1 or more)."""
    data_json = json.dumps(all_data, ensure_ascii=False, default=str)
    return HTML_TEMPLATE.replace('/*__DATA__*/', data_json)


HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>우리집 가계부</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{
  --bg:#0d1117;--card:#161b22;--border:#30363d;
  --text:#e6edf3;--text2:#8b949e;
  --blue:#58a6ff;--green:#3fb950;--red:#f85149;
  --yellow:#d29922;--purple:#bc8cff;--orange:#f0883e;
  --sidebar-w:220px;
}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--text);display:flex;min-height:100vh}
a{color:var(--blue);text-decoration:none}

/* Sidebar */
.sidebar{width:var(--sidebar-w);background:var(--card);border-right:1px solid var(--border);padding:20px 0;position:fixed;height:100vh;overflow-y:auto;z-index:100}
.sidebar .logo{padding:0 20px 24px;font-size:20px;font-weight:700;color:var(--text);border-bottom:1px solid var(--border);margin-bottom:8px}
.sidebar .user-info{padding:12px 20px;font-size:13px;color:var(--text2);border-bottom:1px solid var(--border);margin-bottom:8px}
.sidebar .user-info strong{color:var(--text);font-size:14px;display:block;margin-bottom:2px}
.nav-item{display:flex;align-items:center;gap:10px;padding:10px 20px;color:var(--text2);font-size:14px;cursor:pointer;transition:all .15s;border-left:3px solid transparent}
.nav-item:hover{background:rgba(88,166,255,.08);color:var(--text)}
.nav-item.active{background:rgba(88,166,255,.12);color:var(--blue);border-left-color:var(--blue)}
.nav-item .icon{width:20px;text-align:center;font-size:16px}
.gen-info{padding:16px 20px;font-size:11px;color:var(--text2);position:absolute;bottom:0;border-top:1px solid var(--border);width:100%}

/* Main content */
.content{margin-left:var(--sidebar-w);flex:1;padding:24px 32px;min-height:100vh}
.page{display:none}
.page.active{display:block}
.page-title{font-size:22px;font-weight:700;margin-bottom:20px;display:flex;align-items:center;gap:10px}
.page-title .icon{font-size:24px}
.page-subtitle{color:var(--text2);font-size:13px;margin:-12px 0 20px}

/* KPI Cards */
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px}
.kpi-label{font-size:12px;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;display:block;margin-bottom:6px}
.kpi-value{font-size:24px;font-weight:700;display:block}
.kpi-sub{font-size:12px;color:var(--text2);margin-top:4px;display:block}
.kpi-card.green .kpi-value{color:var(--green)}
.kpi-card.red .kpi-value{color:var(--red)}
.kpi-card.blue .kpi-value{color:var(--blue)}
.kpi-card.purple .kpi-value{color:var(--purple)}
.kpi-card.yellow .kpi-value{color:var(--yellow)}

/* Charts */
.chart-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}
.chart-grid.single{grid-template-columns:1fr}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px}
.chart-card h3{font-size:14px;font-weight:600;margin-bottom:16px;color:var(--text2)}
.chart-card.clickable canvas{cursor:pointer}
.chart-card.clickable h3::after{content:' (클릭시 세부내역)';font-size:10px;color:var(--text2);font-weight:400}
.chart-card canvas{max-height:320px}

/* Tables */
.table-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px;overflow-x:auto;margin-bottom:24px}
.table-card h3{font-size:14px;font-weight:600;margin-bottom:16px;color:var(--text2)}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;padding:10px 12px;border-bottom:2px solid var(--border);color:var(--text2);font-weight:600;white-space:nowrap}
td{padding:8px 12px;border-bottom:1px solid var(--border)}
tr:hover{background:rgba(88,166,255,.04)}
.amount-pos{color:var(--green);font-weight:600}
.amount-neg{color:var(--red);font-weight:600}
.amount-zero{color:var(--text2)}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-income{background:rgba(63,185,80,.15);color:var(--green)}
.badge-expense{background:rgba(248,81,73,.15);color:var(--red)}
.badge-transfer{background:rgba(188,140,255,.15);color:var(--purple)}

/* Filters */
.filter-bar{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;align-items:center}
.filter-bar input,.filter-bar select{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:8px 12px;color:var(--text);font-size:13px;outline:none}
.filter-bar input:focus,.filter-bar select:focus{border-color:var(--blue)}
.filter-bar input{min-width:200px}
.filter-bar select{min-width:120px}

/* Pagination */
.pagination{display:flex;justify-content:center;align-items:center;gap:8px;margin-top:16px}
.pagination button{background:var(--card);border:1px solid var(--border);color:var(--text);padding:6px 12px;border-radius:6px;cursor:pointer;font-size:13px}
.pagination button:hover{border-color:var(--blue)}
.pagination button.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.pagination button:disabled{opacity:.4;cursor:not-allowed}
.pagination .info{font-size:13px;color:var(--text2)}

/* Asset toggles */
.asset-group{margin-bottom:20px}
.asset-group-title{font-size:15px;font-weight:600;margin-bottom:10px;padding:8px 0;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center}
.asset-item{display:flex;justify-content:space-between;align-items:center;padding:10px 14px;border-radius:8px;margin-bottom:4px;transition:background .15s}
.asset-item:hover{background:rgba(255,255,255,.03)}
.asset-item.parent-delegated{opacity:.5;border-left:3px solid var(--purple);background:rgba(188,140,255,.04)}
.asset-item.illiquid:not(.parent-delegated){opacity:.75;border-left:3px solid var(--yellow);background:rgba(210,153,34,.04)}
.asset-item.parent-delegated.illiquid{opacity:.35;border-left:3px solid var(--purple)}
.asset-item .left{display:flex;align-items:center;gap:12px;flex:1;min-width:0}
.asset-item .name{white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:14px}
.asset-item .amount{font-weight:600;font-size:14px;white-space:nowrap}
.toggle-label{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text2);cursor:pointer;white-space:nowrap}
.toggle-label input[type=checkbox]{accent-color:var(--orange);width:16px;height:16px;cursor:pointer}
.toggle-label.parent input[type=checkbox]{accent-color:var(--purple)}
.toggle-label.illiquid input[type=checkbox]{accent-color:var(--yellow)}
.asset-toggles{display:flex;gap:12px}
.asset-legend{display:flex;gap:20px;margin-bottom:16px;padding:12px 16px;background:var(--card);border:1px solid var(--border);border-radius:8px;font-size:12px;color:var(--text2);flex-wrap:wrap}

/* Edit modal */
.edit-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:200;justify-content:center;align-items:center}
.edit-overlay.show{display:flex}
.edit-modal{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:24px;width:480px;max-width:95vw;max-height:90vh;overflow-y:auto}
.edit-modal h3{font-size:16px;margin-bottom:16px}
.edit-modal .field{margin-bottom:12px}
.edit-modal .field label{display:block;font-size:12px;color:var(--text2);margin-bottom:4px}
.edit-modal .field input,.edit-modal .field select{width:100%;background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:8px 10px;color:var(--text);font-size:13px;outline:none}
.edit-modal .field input:focus,.edit-modal .field select:focus{border-color:var(--blue)}
.edit-modal .actions{display:flex;gap:8px;justify-content:flex-end;margin-top:16px}
.btn{padding:8px 16px;border-radius:6px;border:1px solid var(--border);background:var(--card);color:var(--text);font-size:13px;cursor:pointer}
.btn:hover{border-color:var(--blue)}
.btn-primary{background:var(--blue);border-color:var(--blue);color:#fff}
.btn-primary:hover{opacity:.9}
.btn-danger{background:var(--red);border-color:var(--red);color:#fff}
.btn-danger:hover{opacity:.9}
.btn-sm{padding:4px 10px;font-size:11px}
.badge-edited{background:rgba(210,153,34,.2);color:var(--yellow);margin-left:4px}

/* Investment cards */
.inv-card{background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:10px}
.inv-card .inv-name{font-size:14px;font-weight:600;margin-bottom:4px}
.inv-card .inv-company{font-size:12px;color:var(--text2);margin-bottom:10px}
.inv-card .inv-stats{display:flex;gap:20px;flex-wrap:wrap}
.inv-card .inv-stat{font-size:12px}
.inv-card .inv-stat .label{color:var(--text2);display:block}
.inv-card .inv-stat .value{font-weight:600;font-size:14px;margin-top:2px}

/* Insurance cards */
.ins-card{background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px}
.ins-card .ins-left{flex:1;min-width:200px}
.ins-card .ins-name{font-size:14px;font-weight:600;margin-bottom:2px}
.ins-card .ins-company{font-size:12px;color:var(--text2)}
.ins-card .ins-right{text-align:right}
.ins-card .ins-status{display:inline-block;background:rgba(63,185,80,.15);color:var(--green);padding:2px 10px;border-radius:10px;font-size:12px;font-weight:600}
.ins-card .ins-paid{font-size:13px;margin-top:4px;color:var(--text2)}

/* Analysis cards */
.analysis-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:16px;margin-bottom:24px}
.analysis-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px}
.analysis-card h3{font-size:14px;font-weight:600;margin-bottom:12px;color:var(--text2)}
.health-score{font-size:48px;font-weight:700;text-align:center;margin:20px 0}
.health-score.good{color:var(--green)}
.health-score.ok{color:var(--yellow)}
.health-score.bad{color:var(--red)}
.bar-h{display:flex;align-items:center;gap:8px;margin-bottom:8px}
.bar-h .bar-label{width:80px;font-size:12px;color:var(--text2);text-align:right;flex-shrink:0}
.bar-h .bar-track{flex:1;height:20px;background:var(--bg);border-radius:4px;overflow:hidden}
.bar-h .bar-fill{height:100%;border-radius:4px;transition:width .3s}
.bar-h .bar-value{width:80px;font-size:12px;font-weight:600;flex-shrink:0}

/* Responsive */
@media(max-width:1200px){.chart-grid{grid-template-columns:1fr}}
@media(max-width:768px){
  .sidebar{width:60px;overflow:hidden}
  .sidebar .logo,.sidebar .user-info,.sidebar .gen-info,.nav-item span:not(.icon){display:none}
  .nav-item{padding:12px;justify-content:center}
  .content{margin-left:60px;padding:16px}
  .kpi-grid{grid-template-columns:repeat(2,1fr)}
}

/* Person toggle */
.person-toggle{padding:12px 20px;border-bottom:1px solid var(--border)}
.person-toggle .label{font-size:11px;color:var(--text2);margin-bottom:6px}
.person-btns{display:flex;gap:4px;flex-wrap:wrap}
.person-btn{padding:4px 10px;font-size:11px;border-radius:6px;border:1px solid var(--border);background:var(--card);color:var(--text);cursor:pointer}
.person-btn:hover{border-color:var(--blue)}
.person-btn.active{background:var(--blue);border-color:var(--blue);color:#fff}
.drop-zone{border:2px dashed var(--border);border-radius:8px;padding:16px 8px;text-align:center;font-size:11px;color:var(--text2);cursor:pointer;transition:border-color .2s}
.drop-zone:hover,.drop-zone.dragover{border-color:var(--blue);color:var(--blue)}
.file-upload{padding:12px 20px;border-bottom:1px solid var(--border)}
.file-upload .label{font-size:11px;color:var(--text2);margin-bottom:6px}

/* Scrollbar */
::-webkit-scrollbar{width:8px;height:8px}
::-webkit-scrollbar-track{background:var(--bg)}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:4px}
::-webkit-scrollbar-thumb:hover{background:var(--text2)}

/* Print */
@media print{
  .sidebar{display:none}
  .content{margin-left:0}
  .filter-bar,.pagination{display:none}
  body{background:#fff;color:#000}
  .kpi-card,.chart-card,.table-card{border:1px solid #ccc;background:#fff}
}

/* Type filter tabs */
.type-filter{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px;align-items:center}
.type-filter .lbl{font-size:12px;color:var(--text2);margin-right:2px;white-space:nowrap}
.type-btn{padding:4px 12px;font-size:12px;border-radius:20px;border:1px solid var(--border);background:var(--card);color:var(--text2);cursor:pointer;transition:all .15s;white-space:nowrap}
.type-btn:hover{border-color:var(--blue);color:var(--text)}
.type-btn.active{background:var(--blue);border-color:var(--blue);color:#fff}
/* Back button */
.back-btn{display:inline-flex;align-items:center;gap:6px;padding:6px 14px;background:rgba(88,166,255,.1);border:1px solid var(--blue);border-radius:8px;font-size:13px;color:var(--blue);cursor:pointer;margin-bottom:16px;font-weight:500}
.back-btn:hover{background:rgba(88,166,255,.2)}
/* Transaction charts */
.tx-charts{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
@media(max-width:1200px){.tx-charts{grid-template-columns:1fr}}
/* Month summary widget */
.month-compare-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px}
@media(max-width:900px){.month-compare-grid{grid-template-columns:repeat(2,1fr)}}
.compare-kpi{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:12px}
.compare-kpi .lbl{font-size:11px;color:var(--text2);text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px}
.compare-kpi .val{font-size:20px;font-weight:700;display:block}
.compare-kpi .delta{font-size:11px;margin-top:3px;display:block}
.delta.up-good{color:var(--green)}.delta.down-good{color:var(--green)}
.delta.up-bad{color:var(--red)}.delta.down-bad{color:var(--red)}
.delta.neutral{color:var(--text2)}
/* Clear filters button - stand out */
#btn-clear-filters{background:var(--orange);border-color:var(--orange);color:#fff;font-weight:600}
#btn-clear-filters:hover{opacity:.85}
</style>
</head>
<body>

<nav class="sidebar">
  <div class="logo">&#128176; &#44032;&#44228;&#48512;</div>
  <div class="user-info">
    <strong id="user-name"></strong>
    <span id="user-detail"></span>
  </div>
  <a class="nav-item active" data-page="dashboard"><span class="icon">&#128200;</span><span>&#45824;&#49884;&#48372;&#46300;</span></a>
  <a class="nav-item" data-page="cashflow"><span class="icon">&#128178;</span><span>&#54788;&#44552;&#55120;&#47492;</span></a>
  <a class="nav-item" data-page="assets"><span class="icon">&#127974;</span><span>&#51088;&#49328;&#44288;&#47532;</span></a>
  <a class="nav-item" data-page="transactions"><span class="icon">&#128203;</span><span>&#49464;&#48512;&#45236;&#50669;</span></a>
  <a class="nav-item" data-page="investments"><span class="icon">&#128201;</span><span>&#53804;&#51088;&#54788;&#54889;</span></a>
  <a class="nav-item" data-page="insurance"><span class="icon">&#128737;</span><span>&#48372;&#54744;&#54788;&#54889;</span></a>
  <a class="nav-item" data-page="analysis"><span class="icon">&#128161;</span><span>&#48516;&#49437;</span></a>
  <a class="nav-item" data-page="fcf"><span class="icon">&#128184;</span><span>FCF/&#53804;&#51088;&#51312;&#50616;</span></a>
  <div class="person-toggle" id="person-toggle" style="display:none">
    <div class="label">&#45936;&#51060;&#53552; &#48372;&#44592;</div>
    <div class="person-btns" id="person-btns"></div>
  </div>
  <div class="file-upload">
    <div class="label">&#50641;&#49472; &#50629;&#47196;&#46300;</div>
    <div class="drop-zone" id="drop-zone">
      &#128194; &#46300;&#47000;&#44536; &#46608;&#45716; &#53364;&#47533;<br>(.xlsx)
      <input type="file" id="file-input" accept=".xlsx,.xls" multiple style="display:none">
    </div>
  </div>
  <div class="gen-info" id="gen-info"></div>
</nav>

<main class="content">

<!-- Dashboard -->
<div id="page-dashboard" class="page active">
  <div class="page-title"><span class="icon">&#128200;</span> &#51116;&#47924; &#45824;&#49884;&#48372;&#46300;</div>
  <div class="kpi-grid" id="kpi-dashboard"></div>
  <div class="chart-grid">
    <div class="chart-card clickable"><h3>&#50900;&#48324; &#49688;&#51077; vs &#51648;&#52636;</h3><canvas id="chart-monthly"></canvas></div>
    <div class="chart-card clickable">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px">
        <h3 id="exp-pie-title" style="margin:0">&#51648;&#52636; &#52852;&#53580;&#44256;&#47532; &#48708;&#51473;</h3>
        <div id="exp-pie-filter" style="display:flex;gap:4px;align-items:center;flex-wrap:wrap"></div>
      </div>
      <canvas id="chart-category-pie"></canvas>
    </div>
  </div>
  <div class="chart-grid single">
    <div class="chart-card"><h3>&#50900;&#48324; &#51200;&#52629;&#47456; &#52628;&#49464;</h3><canvas id="chart-savings-rate"></canvas></div>
  </div>
  <div class="chart-card">
    <h3 id="month-summary-title">&#52572;&#44540;&#50900; &#54788;&#54889;</h3>
    <div id="month-summary-content"></div>
  </div>
</div>

<!-- Cash Flow -->
<div id="page-cashflow" class="page">
  <div class="page-title"><span class="icon">&#128178;</span> &#54788;&#44552;&#55120;&#47492; &#48516;&#49437;</div>
  <div class="kpi-grid" id="kpi-cashflow"></div>
  <div class="chart-grid single">
    <div class="chart-card clickable"><h3>&#50900;&#48324; &#49688;&#51077;/&#51648;&#52636; &#49345;&#49464;</h3><canvas id="chart-cashflow-detail"></canvas></div>
  </div>
  <div class="chart-grid">
    <div class="chart-card clickable"><h3>&#49688;&#51077; &#44396;&#49457;</h3><canvas id="chart-income-breakdown"></canvas></div>
    <div class="chart-card clickable"><h3>&#51648;&#52636; &#44396;&#49457; (&#50900;&#54217;&#44512;)</h3><canvas id="chart-expense-breakdown"></canvas></div>
  </div>
  <div class="table-card">
    <h3>&#50900;&#48324; &#54788;&#44552;&#55120;&#47492; &#49345;&#49464;</h3>
    <table id="cashflow-table"></table>
  </div>
</div>

<!-- Assets -->
<div id="page-assets" class="page">
  <div class="page-title"><span class="icon">&#127974;</span> &#51088;&#49328;&#44288;&#47532;</div>
  <div class="asset-legend">
    <div style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:2px;background:var(--purple);display:inline-block"></span><span><strong style="color:var(--purple)">&#50948;&#53441;</strong>: &#48512;&#47784;&#45784; &#51088;&#49328;. &#45236; &#52509;&#51088;&#49328;&#50640;&#49436; <strong>&#50756;&#51204;&#55176; &#51228;&#50808;</strong></span></div>
    <div style="display:flex;align-items:center;gap:6px"><span style="width:12px;height:12px;border-radius:2px;background:var(--yellow);display:inline-block"></span><span><strong style="color:var(--yellow)">&#48708;&#50976;&#46041;</strong>: &#45236; &#51088;&#49328;&#50640; <strong>&#54252;&#54632;</strong>&#46104;&#51648;&#47564;, &#50976;&#46041;&#51088;&#49328;&#50640;&#49436;&#45716; &#51228;&#50808; (&#48512;&#46041;&#49328;, &#48372;&#51613;&#44552; &#46321;)</span></div>
  </div>
  <div class="kpi-grid" id="kpi-assets"></div>
  <div class="chart-grid">
    <div class="chart-card"><h3>&#51088;&#49328; &#44396;&#49457; (&#50976;&#46041;&#51088;&#49328;)</h3><canvas id="chart-asset-alloc"></canvas></div>
    <div class="chart-card"><h3>&#53804;&#51088;&#49457; &#51088;&#49328; &#49345;&#49464;</h3><canvas id="chart-invest-alloc"></canvas></div>
  </div>
  <div id="asset-list"></div>
</div>

<!-- Transactions -->
<div id="page-transactions" class="page">
  <div class="page-title"><span class="icon">&#128203;</span> &#49464;&#48512;&#45236;&#50669;</div>
  <button id="tx-back-btn" class="back-btn" style="display:none" onclick="goBackFromDrill()">&#8592; <span id="tx-back-label">&#51060;&#51204; &#54168;&#51060;&#51648;&#47196;</span></button>
  <div class="filter-bar">
    <input type="text" id="tx-search" placeholder="&#44160;&#49353;&#50612;&#47484; &#51077;&#47141;&#54616;&#49464;&#50836;...">
    <select id="tx-type"><option value="">&#47784;&#46304; &#53440;&#51077;</option><option value="&#51648;&#52636;">&#51648;&#52636;</option><option value="&#49688;&#51077;">&#49688;&#51077;</option><option value="&#51060;&#52404;">&#51060;&#52404;</option></select>
    <select id="tx-category"><option value="">&#47784;&#46304; &#52852;&#53580;&#44256;&#47532;</option></select>
    <select id="tx-month"><option value="">&#47784;&#46304; &#50900;</option></select>
    <span class="info" id="tx-count"></span>
    <button class="btn btn-sm" id="btn-clear-filters" onclick="clearFilters()">&#54596;&#53552; &#52488;&#44592;&#54868;</button>
    <button class="btn btn-sm" onclick="exportTxCsv()">CSV &#45796;&#50868;&#47196;&#46300;</button>
    <button class="btn btn-sm" onclick="resetAllEdits()" id="btn-reset-edits" style="display:none">&#49688;&#51221; &#52488;&#44592;&#54868;</button>
  </div>
  <div id="tx-charts-container" style="display:none">
    <div class="tx-charts">
      <div class="chart-card clickable"><h3 id="tx-chart-exp-title">&#52852;&#53580;&#44256;&#47532;&#48324; &#51648;&#52636;</h3><canvas id="chart-tx-expense"></canvas></div>
      <div class="chart-card"><h3 id="tx-chart-mon-title">&#50900;&#48324; &#49688;&#51077;/&#51648;&#52636;</h3><canvas id="chart-tx-monthly"></canvas></div>
    </div>
  </div>
  <div class="table-card">
    <table id="tx-table"></table>
  </div>
  <div class="pagination" id="tx-pagination"></div>
</div>

<!-- Edit Modal -->
<div class="edit-overlay" id="edit-overlay" onclick="if(event.target===this)closeEditModal()">
  <div class="edit-modal">
    <h3 id="edit-title">&#44144;&#47000; &#49688;&#51221;</h3>
    <input type="hidden" id="edit-idx">
    <div class="field"><label>&#45216;&#51676;</label><input type="date" id="edit-date"></div>
    <div class="field"><label>&#53440;&#51077;</label><select id="edit-type"><option>&#51648;&#52636;</option><option>&#49688;&#51077;</option><option>&#51060;&#52404;</option></select></div>
    <div class="field"><label>&#45824;&#48516;&#47448;</label><input type="text" id="edit-category" list="cat-list"></div>
    <datalist id="cat-list"></datalist>
    <div class="field"><label>&#49548;&#48516;&#47448;</label><input type="text" id="edit-subcategory"></div>
    <div class="field"><label>&#45236;&#50857;</label><input type="text" id="edit-description"></div>
    <div class="field"><label>&#44552;&#50529;</label><input type="number" id="edit-amount"></div>
    <div class="actions">
      <button class="btn" onclick="closeEditModal()">&#52712;&#49548;</button>
      <button class="btn btn-danger btn-sm" id="btn-revert" onclick="revertEdit()" style="display:none">&#50896;&#47000;&#45824;&#47196;</button>
      <button class="btn btn-primary" onclick="saveEdit()">&#51200;&#51109;</button>
    </div>
  </div>
</div>

<!-- Confirm Modal -->
<div class="edit-overlay" id="confirm-overlay" onclick="if(event.target===this)closeConfirmModal()">
  <div class="edit-modal" style="max-width:360px">
    <h3>&#54869;&#51064;</h3>
    <p id="confirm-msg" style="color:var(--text2);font-size:14px;margin:12px 0 20px;line-height:1.6"></p>
    <div class="actions">
      <button class="btn" onclick="closeConfirmModal()">&#52712;&#49548;</button>
      <button class="btn btn-danger" onclick="if(_confirmOkFn){closeConfirmModal();_confirmOkFn();}">&#54869;&#51064;</button>
    </div>
  </div>
</div>

<!-- Investments -->
<div id="page-investments" class="page">
  <div class="page-title"><span class="icon">&#128201;</span> &#53804;&#51088;&#54788;&#54889;</div>
  <div class="kpi-grid" id="kpi-investments"></div>
  <div class="chart-grid">
    <div class="chart-card"><h3>&#54252;&#53944;&#54260;&#47532;&#50724; &#48176;&#48516;</h3><canvas id="chart-portfolio"></canvas></div>
    <div class="chart-card"><h3>&#49688;&#51061;&#47456;</h3><canvas id="chart-returns"></canvas></div>
  </div>
  <div id="inv-type-filter" class="type-filter"></div>
  <div id="investment-list"></div>
</div>

<!-- Insurance -->
<div id="page-insurance" class="page">
  <div class="page-title"><span class="icon">&#128737;</span> &#48372;&#54744;&#54788;&#54889;</div>
  <div class="kpi-grid" id="kpi-insurance"></div>
  <div id="insurance-list"></div>
</div>

<!-- Analysis -->
<div id="page-analysis" class="page">
  <div class="page-title"><span class="icon">&#128161;</span> &#51116;&#47924; &#48516;&#49437; &#48143; &#51064;&#49324;&#51060;&#53944;</div>
  <div class="analysis-grid" id="analysis-content"></div>
  <div class="chart-grid">
    <div class="chart-card"><h3>&#52852;&#53580;&#44256;&#47532;&#48324; &#50900;&#48324; &#52628;&#49464;</h3><canvas id="chart-category-trend"></canvas></div>
    <div class="chart-card"><h3>&#50836;&#51068;&#48324; &#51648;&#52636; &#54056;&#53556;</h3><canvas id="chart-day-pattern"></canvas></div>
  </div>
  <div class="table-card">
    <h3>&#51060;&#49345; &#51648;&#52636; &#44048;&#51648; (&#50900;&#54217;&#44512; &#45824;&#48708; 2&#48176; &#51060;&#49345;)</h3>
    <table id="anomaly-table"></table>
  </div>
  <div style="margin-bottom:16px;padding:12px 16px;background:var(--card);border:1px solid var(--border);border-radius:8px;display:flex;gap:10px;align-items:center;flex-wrap:wrap">
    <span style="font-size:12px;color:var(--text2);white-space:nowrap">&#uae30&#uac04:</span>
    <div id="analysis-chart-filter" style="display:flex;gap:4px;flex-wrap:wrap"></div>
  </div>
  <div class="chart-grid">
    <div class="chart-card"><h3>&#44208;&#51228;&#49688;&#45800;&#48324; &#51648;&#52636;</h3><canvas id="chart-payment-method"></canvas></div>
    <div class="chart-card"><h3>&#49884;&#44036;&#45824;&#48324; &#51648;&#52636; &#54056;&#53556;</h3><canvas id="chart-hour-pattern"></canvas></div>
  </div>
</div>

<!-- FCF -->
<div id="page-fcf" class="page">
  <div class="page-title"><span class="icon">&#128184;</span> FCF &#48516;&#49437; &#48143; &#53804;&#51088; &#51312;&#50616;</div>
  <p class="page-subtitle">Free Cash Flow = &#49688;&#51077; - &#51648;&#52636;. &#44144;&#47000; &#49688;&#51221;&#51060; &#48152;&#50689;&#46121;&#45768;&#45796;.</p>
  <div class="kpi-grid" id="kpi-fcf"></div>
  <div class="chart-grid">
    <div class="chart-card clickable"><h3>&#50900;&#48324; FCF</h3><canvas id="chart-fcf-monthly"></canvas></div>
    <div class="chart-card"><h3>&#45572;&#51201; FCF</h3><canvas id="chart-fcf-cumulative"></canvas></div>
  </div>
  <div class="analysis-grid" id="fcf-advice"></div>
</div>

</main>

<script>
const ALL_DATA = /*__DATA__*/;
const PERSONS = Array.isArray(ALL_DATA) ? ALL_DATA : [ALL_DATA];
let currentPerson = 0;
let DATA = PERSONS[0];

// ===== Utilities =====
const fmt = (n) => {
  if (n === 0) return '-';
  const abs = Math.abs(Math.round(n));
  if (abs >= 100000000) return (n < 0 ? '-' : '') + (abs / 100000000).toFixed(1) + '\uc5b5';
  if (abs >= 10000) return (n < 0 ? '-' : '') + Math.round(abs / 10000).toLocaleString() + '\ub9cc';
  return Math.round(n).toLocaleString() + '\uc6d0';
};
const fmtFull = (n) => Math.round(n).toLocaleString() + '\uc6d0';
const fmtPct = (n) => (n >= 0 ? '+' : '') + n.toFixed(1) + '%';
const amtClass = (n) => n > 0 ? 'amount-pos' : n < 0 ? 'amount-neg' : 'amount-zero';
const badgeClass = (t) => t === '\uc218\uc785' ? 'badge-income' : t === '\uc9c0\ucd9c' ? 'badge-expense' : 'badge-transfer';

const COLORS = ['#58a6ff','#3fb950','#f85149','#d29922','#bc8cff','#f0883e','#79c0ff','#56d364','#ff7b72','#e3b341','#d2a8ff','#ffa657','#a5d6ff','#7ee787','#ffa198'];

// ===== Shared chart plugins =====
const pieLblPlugin = { id:'pieLbl', afterDatasetsDraw(chart) { const ctx=chart.ctx,ds=chart.data.datasets[0],tot=ds.data.reduce((a,b)=>a+b,0); chart.getDatasetMeta(0).data.forEach((arc,j)=>{ const v=ds.data[j]; if(!v||v/tot<0.05)return; const mid=arc.startAngle+(arc.endAngle-arc.startAngle)/2,r=(arc.innerRadius+arc.outerRadius)/2,x=arc.x+Math.cos(mid)*r,y=arc.y+Math.sin(mid)*r; ctx.save();ctx.fillStyle='#fff';ctx.font='bold 11px system-ui';ctx.textAlign='center';ctx.textBaseline='middle';ctx.fillText(fmt(v),x,y);ctx.restore(); }); } };
const barFmtPlugin = { id:'barFmt', afterDatasetsDraw(chart) { const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,j)=>{ const v=chart.data.datasets[0].data[j]; if(v===null||v===undefined||v===0)return; const isNeg=v<0; ctx.save();ctx.fillStyle='#e6edf3';ctx.font='bold 10px system-ui';ctx.textAlign='center';ctx.textBaseline=isNeg?'top':'bottom';ctx.fillText(fmt(v),bar.x,isNeg?bar.y+3:bar.y-3);ctx.restore(); }); } };
const hBarFmtPlugin = { id:'hBarFmt', afterDatasetsDraw(chart) { const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,j)=>{ const v=chart.data.datasets[0].data[j]; if(!v)return; const isNeg=v<0; ctx.save();ctx.fillStyle='#e6edf3';ctx.font='bold 10px system-ui';ctx.textBaseline='middle';ctx.textAlign=isNeg?'right':'left';ctx.fillText(fmt(Math.abs(v)),isNeg?bar.x-4:bar.x+4,bar.y);ctx.restore(); }); } };
const pctBarPlugin = { id:'pctBar', afterDatasetsDraw(chart) { const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,j)=>{ const v=chart.data.datasets[0].data[j]; if(v===null||v===undefined)return; const isNeg=v<0; ctx.save();ctx.fillStyle='#e6edf3';ctx.font='bold 10px system-ui';ctx.textBaseline='middle';ctx.textAlign=isNeg?'right':'left';ctx.fillText(v.toFixed(1)+'%',isNeg?bar.x-4:bar.x+4,bar.y);ctx.restore(); }); } };
const lineLblPlugin = { id:'lineLbl', afterDatasetsDraw(chart) { const ctx=chart.ctx; chart.data.datasets.forEach((ds,di)=>{ chart.getDatasetMeta(di).data.forEach((pt,j)=>{ const v=ds.data[j]; if(!v)return; ctx.save();ctx.fillStyle='#8b949e';ctx.font='9px system-ui';ctx.textAlign='center';ctx.textBaseline='bottom';ctx.fillText(fmt(v),pt.x,pt.y-6);ctx.restore(); }); }); } };

let charts = {};
const destroyChart = (id) => { if (charts[id]) { charts[id].destroy(); delete charts[id]; } };

let drillSource = '';
let currentInvType = '';
let expPieFilter = '';
let analysisChartFilter = '';
let _confirmOkFn = null;

// ===== State =====
const lsKey = (base) => base + (PERSONS.length > 1 ? '_' + (currentPerson === -1 ? 'merged' : 'p' + currentPerson) : '');
let parentAssets = {}, illiquidAssets = {}, excludedInv = {}, txEdits = {};
function loadState() {
  parentAssets = JSON.parse(localStorage.getItem(lsKey('parentAssets')) || '{}');
  illiquidAssets = JSON.parse(localStorage.getItem(lsKey('illiquidAssets')) || '{}');
  excludedInv = JSON.parse(localStorage.getItem(lsKey('excludedInv')) || '{}');
  txEdits = JSON.parse(localStorage.getItem(lsKey('txEdits')) || '{}');
}
loadState();

const isParent = (id) => !!parentAssets[id];
const isIlliquid = (id) => !!illiquidAssets[id];
const isExcludedInv = (idx) => !!excludedInv[idx];

const toggleParent = (id) => {
  if (parentAssets[id]) delete parentAssets[id]; else parentAssets[id] = true;
  localStorage.setItem(lsKey('parentAssets'), JSON.stringify(parentAssets));
  renderAssets(); renderDashboard();
};
const toggleIlliquid = (id) => {
  if (illiquidAssets[id]) delete illiquidAssets[id]; else illiquidAssets[id] = true;
  localStorage.setItem(lsKey('illiquidAssets'), JSON.stringify(illiquidAssets));
  renderAssets(); renderDashboard();
};
const toggleExcludedInv = (idx) => {
  if (excludedInv[idx]) delete excludedInv[idx]; else excludedInv[idx] = true;
  localStorage.setItem(lsKey('excludedInv'), JSON.stringify(excludedInv));
  renderInvestments();
};

// Get transaction with edits applied
function getTx(i) {
  const t = DATA.transactions[i];
  const e = txEdits[i];
  return e ? {...t, ...e, _edited: true} : t;
}
function getAllTx() { return DATA.transactions.map((_, i) => ({...getTx(i), _idx: i})); }

// ===== Computed Data =====
function _isIncomeTx(type, category) {
  if (type === '\uc218\uc785') return true;
  if (type === '\uc9c0\ucd9c') return false;
  if (DATA.income[category] && !DATA.expense[category]) return true;
  if (DATA.expense[category] && !DATA.income[category]) return false;
  return false;
}
function getEditedCashFlow() {
  const income = {};
  for (const [k, v] of Object.entries(DATA.income)) income[k] = [...v];
  const expense = {};
  for (const [k, v] of Object.entries(DATA.expense)) expense[k] = [...v];
  const N = DATA.months.length;
  for (const [idx, edits] of Object.entries(txEdits)) {
    const orig = DATA.transactions[parseInt(idx)];
    if (!orig) continue;
    const mi = DATA.months.indexOf(orig.date.substring(0, 7));
    if (mi < 0) continue;
    const ed = {...orig, ...edits};
    const typeChanged = orig.type !== ed.type || orig.category !== ed.category;
    const amtChanged = orig.amount !== ed.amount;
    if (!typeChanged && !amtChanged) continue;
    const origIsInc = _isIncomeTx(orig.type, orig.category);
    const edIsInc = _isIncomeTx(ed.type, ed.category);
    if (typeChanged) {
      if (origIsInc && income[orig.category]) income[orig.category][mi] = Math.max(0, (income[orig.category][mi] || 0) - orig.amount);
      if (!origIsInc && expense[orig.category]) expense[orig.category][mi] = Math.max(0, (expense[orig.category][mi] || 0) - orig.amount);
      if (edIsInc) { if (!income[ed.category]) income[ed.category] = new Array(N).fill(0); income[ed.category][mi] += parseFloat(ed.amount) || 0; }
      else { if (!expense[ed.category]) expense[ed.category] = new Array(N).fill(0); expense[ed.category][mi] += parseFloat(ed.amount) || 0; }
    } else if (amtChanged) {
      const diff = (parseFloat(ed.amount) || 0) - orig.amount;
      if (origIsInc) { if (income[orig.category]) income[orig.category][mi] += diff; }
      else { if (expense[orig.category]) expense[orig.category][mi] += diff; }
    }
  }
  return { income, expense };
}
function getMonthlyIncome() {
  const { income } = getEditedCashFlow();
  return DATA.months.map((_, i) => Object.values(income).reduce((s, v) => s + (v[i] || 0), 0));
}
function getMonthlyExpense() {
  const { expense } = getEditedCashFlow();
  return DATA.months.map((_, i) => Object.values(expense).reduce((s, v) => s + (v[i] || 0), 0));
}
function getTotalAssets(excludeParent = true, excludeIlliquid = false) {
  return DATA.assets.reduce((s, a) => {
    if (excludeParent && isParent(a.id)) return s;
    if (excludeIlliquid && isIlliquid(a.id)) return s;
    return s + a.amount;
  }, 0);
}
function getAvgExpenseByCategory() {
  const { expense } = getEditedCashFlow();
  const result = {};
  for (const [cat, vals] of Object.entries(expense)) {
    const nonZero = vals.filter(v => v > 0);
    result[cat] = nonZero.length ? nonZero.reduce((a, b) => a + b, 0) / nonZero.length : 0;
  }
  return result;
}

// ===== Multi-Person =====
function getMergedData() {
  if (PERSONS.length === 1) return PERSONS[0];
  const allMonths = [...new Set(PERSONS.flatMap(p => p.months))].sort();
  const mergeFlow = (key) => {
    const result = {};
    PERSONS.forEach(p => {
      for (const [cat, vals] of Object.entries(p[key])) {
        if (!result[cat]) result[cat] = new Array(allMonths.length).fill(0);
        vals.forEach((v, i) => {
          const mi = allMonths.indexOf(p.months[i]);
          if (mi >= 0) result[cat][mi] += v;
        });
      }
    });
    return result;
  };
  return {
    customer: { ...PERSONS[0].customer, name: PERSONS.map(p => p.customer.name).join(' & ') },
    months: allMonths,
    income: mergeFlow('income'),
    expense: mergeFlow('expense'),
    assets: PERSONS.flatMap((p, pi) => p.assets.map(a => ({...a, id: 'p' + pi + '-' + a.id}))),
    insurance: PERSONS.flatMap(p => p.insurance),
    investments: PERSONS.flatMap(p => p.investments),
    transactions: PERSONS.flatMap(p => p.transactions).sort((a, b) => b.date.localeCompare(a.date)),
    generatedAt: PERSONS[0].generatedAt,
  };
}

function switchPerson(idx) {
  currentPerson = idx;
  DATA = idx === -1 ? getMergedData() : PERSONS[idx];
  loadState();
  currentInvType = '';
  drillSource = '';
  expPieFilter = '';
  analysisChartFilter = '';
  document.querySelectorAll('.person-btn').forEach(b => b.classList.toggle('active', parseInt(b.dataset.person) === idx));
  document.getElementById('user-name').textContent = DATA.customer.name;
  document.getElementById('user-detail').textContent = DATA.customer.gender + ' / ' + DATA.customer.age + '\uc138';
  const activePage = document.querySelector('.nav-item.active');
  if (activePage) initPage(activePage.dataset.page);
}

function updatePersonToggle() {
  const toggle = document.getElementById('person-toggle');
  const btns = document.getElementById('person-btns');
  if (PERSONS.length > 1) {
    toggle.style.display = '';
    const labels = PERSONS.map((p, i) => p.customer.name || ('Person ' + (i + 1)));
    btns.innerHTML = labels.map((l, i) => '<button class="btn btn-sm person-btn ' + (i === currentPerson ? 'active' : '') + '" data-person="' + i + '" onclick="switchPerson(' + i + ')">' + l + '</button>').join('') +
      '<button class="btn btn-sm person-btn ' + (currentPerson === -1 ? 'active' : '') + '" data-person="-1" onclick="switchPerson(-1)">\ud1b5\ud569</button>';
  } else {
    toggle.style.display = 'none';
  }
}

function parseExcelClient(arrayBuffer) {
  const wb = XLSX.read(arrayBuffer, { type: 'array', cellDates: true });
  const ws1 = wb.Sheets['\ubf45\uc0d0\ud604\ud669'];
  const ws2 = wb.Sheets['\uac00\uacc4\ubd80 \ub0b4\uc5ed'];
  if (!ws1 || !ws2) throw new Error('\ubf45\ud06c\uc0d0\ub7ec\ub4dc \ud615\uc2dd\uc774 \uc544\ub2d9\ub2c8\ub2e4');
  const cell = (ws, r, c) => { const addr = XLSX.utils.encode_cell({r: r-1, c: c-1}); const cl = ws[addr]; return cl ? cl.v : null; };
  const cellDate = (ws, r, c) => { const addr = XLSX.utils.encode_cell({r: r-1, c: c-1}); const cl = ws[addr]; if (!cl) return ''; if (cl.t === 'd') return cl.v.toISOString().substring(0,10); if (typeof cl.v === 'number') { const d = XLSX.SSF.parse_date_code(cl.v); return d.y + '-' + String(d.m).padStart(2,'0') + '-' + String(d.d).padStart(2,'0'); } return String(cl.v).substring(0,10); };
  const customer = { name: cell(ws1,6,2)||'', gender: cell(ws1,6,3)||'', age: cell(ws1,6,4)||0, creditScore: cell(ws1,6,5)||0 };
  const months = []; for (let c=5;c<=17;c++){const v=cell(ws1,11,c);if(v)months.push(String(v));}
  const income = {}; for (let r=12;r<=15;r++){const cat=cell(ws1,r,2);if(cat&&cat!=='\uc6d4\uc218\uc785 \ucd1d\uacc4'){income[cat]=months.map((_,i)=>cell(ws1,r,5+i)||0);}}
  const expense = {}; for (let r=17;r<=34;r++){const cat=cell(ws1,r,2);if(cat&&cat!=='\uc6d4\uc9c0\ucd9c \ucd1d\uacc4'&&cat!=='\uc21c\uc218\uc785 \ucd1d\uacc4'){expense[cat]=months.map((_,i)=>cell(ws1,r,5+i)||0);}}
  const assets = []; let curCat = '';
  for (let r=43;r<=84;r++){const b=cell(ws1,r,2),c=cell(ws1,r,3),e=cell(ws1,r,5);if(b&&!c){curCat=b;continue;}if(c){if(b)curCat=b;let amt=e||0;if(typeof amt==='string')amt=0;assets.push({id:'asset-'+r,category:curCat,name:c,amount:parseFloat(amt)||0});}}
  const insurance = []; for(let r=94;r<=97;r++){const nm=cell(ws1,r,3);if(!nm)continue;insurance.push({company:cell(ws1,r,2)||'',name:nm,status:cell(ws1,r,5)||'',totalPaid:cell(ws1,r,6)||0,startDate:cellDate(ws1,r,7),endDate:cellDate(ws1,r,8)});}
  const investments = []; for(let r=104;r<=112;r++){const nm=cell(ws1,r,4);if(!nm)continue;investments.push({type:cell(ws1,r,2)||'',company:cell(ws1,r,3)||'',name:nm,principal:parseFloat(cell(ws1,r,6)||0),currentValue:parseFloat(cell(ws1,r,7)||0),returnRate:parseFloat(cell(ws1,r,8)||0),startDate:cellDate(ws1,r,9),endDate:cellDate(ws1,r,10)});}
  const transactions = []; const range = XLSX.utils.decode_range(ws2['!ref']||'A1');
  for(let r=2;r<=range.e.r+1;r++){const d=cell(ws2,r,1);if(!d)continue;const t=cell(ws2,r,2);transactions.push({date:(d instanceof Date)?d.toISOString().substring(0,10):String(d).substring(0,10),time:t?String(t):'',type:cell(ws2,r,3)||'',category:cell(ws2,r,4)||'',subcategory:cell(ws2,r,5)||'',description:cell(ws2,r,6)||'',amount:parseFloat(cell(ws2,r,7)||0),currency:cell(ws2,r,8)||'KRW',paymentMethod:cell(ws2,r,9)||'',memo:cell(ws2,r,10)||''});}
  return {customer,months,income,expense,assets,insurance,investments,transactions,generatedAt:new Date().toLocaleString()};
}

function handleFileUpload(files) {
  const xlsxFiles = [...files].filter(f => f.name.match(/\.xlsx?$/i));
  if (!xlsxFiles.length) return;
  Promise.all(xlsxFiles.map(f => f.arrayBuffer().then(buf => parseExcelClient(buf)))).then(results => {
    PERSONS.length = 0;
    results.forEach(d => PERSONS.push(d));
    currentPerson = 0;
    DATA = PERSONS[0];
    loadState();
    updatePersonToggle();
    document.getElementById('user-name').textContent = DATA.customer.name;
    document.getElementById('user-detail').textContent = DATA.customer.gender + ' / ' + DATA.customer.age + '\uc138';
    document.getElementById('gen-info').textContent = '\uc0dd\uc131: ' + DATA.generatedAt;
    // Reset cached dropdowns
    document.getElementById('tx-category').innerHTML = '<option value="">\ubaa8\ub4e0 \uce74\ud14c\uace0\ub9ac</option>';
    document.getElementById('tx-month').innerHTML = '<option value="">\ubaa8\ub4e0 \uc6d4</option>';
    renderDashboard();
    document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
    document.querySelector('.nav-item[data-page="dashboard"]').classList.add('active');
    document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
    document.getElementById('page-dashboard').classList.add('active');
  }).catch(err => alert('\ud30c\uc77c \uc77d\uae30 \uc624\ub958: ' + err.message));
}

// ===== Navigation =====
document.querySelectorAll('.nav-item').forEach(el => {
  el.addEventListener('click', (e) => {
    e.preventDefault();
    drillSource = '';
    document.getElementById('tx-back-btn').style.display = 'none';
    const page = el.dataset.page;
    document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
    el.classList.add('active');
    document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
    document.getElementById('page-' + page).classList.add('active');
    initPage(page);
  });
});

function initPage(page) {
  switch (page) {
    case 'dashboard': renderDashboard(); break;
    case 'cashflow': renderCashflow(); break;
    case 'assets': renderAssets(); break;
    case 'transactions': renderTransactions(); break;
    case 'investments': renderInvestments(); break;
    case 'insurance': renderInsurance(); break;
    case 'analysis': renderAnalysis(); break;
    case 'fcf': renderFcf(); break;
  }
}

// ===== Drill-Down =====
function drillDown(opts = {}) {
  const srcEl = document.querySelector('.nav-item.active');
  drillSource = srcEl ? srcEl.dataset.page : '';
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  document.querySelector('.nav-item[data-page="transactions"]').classList.add('active');
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.getElementById('page-transactions').classList.add('active');
  const backBtn = document.getElementById('tx-back-btn');
  if (drillSource && drillSource !== 'transactions') {
    const labels = {cashflow:'\ud604\uae08\ud750\ub984\uc73c\ub85c', dashboard:'\ub300\uc2dc\ubcf4\ub4dc\ub85c', fcf:'FCF\ub85c', analysis:'\ubd84\uc11d\uc73c\ub85c'};
    document.getElementById('tx-back-label').textContent = labels[drillSource] || '\uc774\uc804 \ud398\uc774\uc9c0\ub85c';
    backBtn.style.display = '';
  } else { backBtn.style.display = 'none'; }
  // Reset category dropdown so it gets repopulated
  const catSel = document.getElementById('tx-category');
  catSel.innerHTML = '<option value="">\ubaa8\ub4e0 \uce74\ud14c\uace0\ub9ac</option>';
  document.getElementById('tx-month').innerHTML = '<option value="">\ubaa8\ub4e0 \uc6d4</option>';
  renderTransactions();
  document.getElementById('tx-search').value = opts.search || '';
  document.getElementById('tx-type').value = opts.type || '';
  if (opts.category) {
    const sel = document.getElementById('tx-category');
    if (![...sel.options].some(o => o.value === opts.category)) { const o = document.createElement('option'); o.value = opts.category; o.text = opts.category; sel.add(o); }
    sel.value = opts.category;
  } else { document.getElementById('tx-category').value = ''; }
  document.getElementById('tx-month').value = opts.month || '';
  filterTx();
}

function goBackFromDrill() {
  const src = drillSource; drillSource = '';
  document.getElementById('tx-back-btn').style.display = 'none';
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  const target = document.querySelector('.nav-item[data-page="' + src + '"]');
  if (target) { target.classList.add('active'); document.querySelectorAll('.page').forEach(p => p.classList.remove('active')); document.getElementById('page-' + src).classList.add('active'); initPage(src); }
}

function clearFilters() {
  document.getElementById('tx-search').value = '';
  document.getElementById('tx-type').value = '';
  document.getElementById('tx-category').value = '';
  document.getElementById('tx-month').value = '';
  txPage = 1;
  filterTx();
}

// ===== Confirm Modal =====
function showConfirm(msg, okFn) {
  document.getElementById('confirm-msg').textContent = msg;
  _confirmOkFn = okFn;
  document.getElementById('confirm-overlay').classList.add('show');
}
function closeConfirmModal() {
  document.getElementById('confirm-overlay').classList.remove('show');
  _confirmOkFn = null;
}

// ===== Expense Pie Filter =====
function getExpByCategoryFiltered(filter) {
  const { expense } = getEditedCashFlow();
  const N = DATA.months.length;
  let indices;
  if (!filter) {
    indices = Array.from({length: N}, (_, i) => i);
  } else if (filter === '3' || filter === '6') {
    const cnt = Math.min(parseInt(filter), N);
    indices = Array.from({length: cnt}, (_, i) => N - cnt + i);
  } else {
    const mi = DATA.months.indexOf(filter);
    indices = mi >= 0 ? [mi] : [];
  }
  const result = {};
  for (const [cat, vals] of Object.entries(expense)) {
    const subVals = indices.map(i => vals[i] || 0);
    const nonZero = subVals.filter(v => v > 0);
    if (nonZero.length) result[cat] = nonZero.reduce((a, b) => a + b, 0) / nonZero.length;
  }
  return result;
}
function buildExpPieFilterHtml(f) {
  const expMonths = DATA.months.slice().reverse();
  return '<button class="type-btn' + (f===''?' active':'') + '" onclick="setExpPieFilter(\'\')">\uc804\uccb4</button>' +
    '<button class="type-btn' + (f==='6'?' active':'') + '" onclick="setExpPieFilter(\'6\')">6\uac1c\uc6d4</button>' +
    '<button class="type-btn' + (f==='3'?' active':'') + '" onclick="setExpPieFilter(\'3\')">3\uac1c\uc6d4</button>' +
    '<select style="background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:2px 6px;color:var(--text);font-size:11px" onchange="setExpPieFilter(this.value)">' +
    '<option value="">\uc6d4 \uc120\ud0dd</option>' +
    expMonths.map(m => '<option value="' + m + '"' + (f===m?' selected':'') + '>' + m + '</option>').join('') +
    '</select>';
}
function setExpPieFilter(f) {
  expPieFilter = f;
  const filterDiv = document.getElementById('exp-pie-filter');
  if (filterDiv) filterDiv.innerHTML = buildExpPieFilterHtml(f);
  renderExpPieOnly();
}
function buildAnalysisChartFilterHtml(f) {
  const months = DATA.months.slice().reverse();
  return '<button class="type-btn' + (f===''?' active':'') + '" onclick="setAnalysisChartFilter(\'\')">\uc804\uccb4</button>' +
    '<button class="type-btn' + (f==='1'?' active':'') + '" onclick="setAnalysisChartFilter(\'1\')">\ucd5c\uadfc 1\uac1c\uc6d4</button>' +
    '<button class="type-btn' + (f==='3'?' active':'') + '" onclick="setAnalysisChartFilter(\'3\')">\ucd5c\uadfc 3\uac1c\uc6d4</button>' +
    '<button class="type-btn' + (f==='6'?' active':'') + '" onclick="setAnalysisChartFilter(\'6\')">\uc0c1\ubc18\uae30</button>' +
    '<select style="background:var(--bg);border:1px solid var(--border);border-radius:6px;padding:2px 6px;color:var(--text);font-size:11px" onchange="setAnalysisChartFilter(this.value)">' +
    '<option value="">\uc6d4 \uc120\ud0dd</option>' +
    months.map(m => '<option value="' + m + '"' + (f===m?' selected':'') + '>' + m + '</option>').join('') +
    '</select>';
}
function setAnalysisChartFilter(f) {
  analysisChartFilter = f;
  const filterDiv = document.getElementById('analysis-chart-filter');
  if (filterDiv) filterDiv.innerHTML = buildAnalysisChartFilterHtml(f);
  renderAnalysisChartsOnly();
}
function getAnalysisChartTxRange() {
  const N = DATA.months.length;
  let indices;
  if (!analysisChartFilter) {
    indices = Array.from({length: N}, (_, i) => i);
  } else if (analysisChartFilter === '3' || analysisChartFilter === '6' || analysisChartFilter === '1') {
    const cnt = Math.min(parseInt(analysisChartFilter), N);
    indices = Array.from({length: cnt}, (_, i) => N - cnt + i);
  } else {
    const mi = DATA.months.indexOf(analysisChartFilter);
    indices = mi >= 0 ? [mi] : [];
  }
  const startDate = indices.length ? DATA.months[indices[0]] : '';
  const endDate = indices.length ? DATA.months[indices[indices.length-1]] : '';
  return {indices, startDate, endDate};
}
function renderExpPieOnly() {
  const filter = expPieFilter;
  const titleEl = document.getElementById('exp-pie-title');
  if (titleEl) {
    const suffix = filter==='' ? '' : filter==='3' ? ' (\ucd5c\uadfc 3\uac1c\uc6d4)' : filter==='6' ? ' (\ucd5c\uadfc 6\uac1c\uc6d4)' : ' (' + filter + ')';
    titleEl.textContent = '\uc9c0\ucd9c \uce74\ud14c\uace0\ub9ac \ube44\uc911' + suffix;
  }
  const data = getExpByCategoryFiltered(filter);
  const sorted = Object.entries(data).filter(([,v]) => v > 0).sort((a,b) => b[1]-a[1]).slice(0, 10);
  destroyChart('chart-category-pie');
  if (!sorted.length) return;
  charts['chart-category-pie'] = new Chart(document.getElementById('chart-category-pie'), {
    type: 'doughnut',
    data: { labels: sorted.map(([k])=>k), datasets: [{ data: sorted.map(([,v])=>Math.round(v)), backgroundColor: COLORS.slice(0,sorted.length) }] },
    options: { responsive: true, onClick: (e,els) => { if(els.length) { const mp = filter&&filter.length===7?{month:filter}:{}; drillDown({ type:'\uc9c0\ucd9c', category:sorted[els[0].index][0], ...mp }); } }, plugins: { legend: { position:'right', labels: { color:'#8b949e', padding:8, font:{size:11} } } } },
    plugins: [pieLblPlugin]
  });
}

// ===== Dashboard =====
function renderDashboard() {
  const inc = getMonthlyIncome();
  const exp = getMonthlyExpense();
  const totalAsset = getTotalAssets(true, false);
  const liquidAsset = getTotalAssets(true, true);
  const parentTotal = getTotalAssets(false, false) - totalAsset;
  const illiquidTotal = totalAsset - liquidAsset;
  const avgInc = inc.filter(v => v > 0).reduce((a, b) => a + b, 0) / Math.max(inc.filter(v => v > 0).length, 1);
  const avgExp = exp.filter(v => v > 0).reduce((a, b) => a + b, 0) / Math.max(exp.filter(v => v > 0).length, 1);
  const savingsRate = avgInc > 0 ? ((avgInc - avgExp) / avgInc * 100) : 0;

  document.getElementById('kpi-dashboard').innerHTML = `
    <div class="kpi-card blue"><span class="kpi-label">\ub0b4 \ucd1d\uc790\uc0b0</span><span class="kpi-value">${fmt(totalAsset)}</span><span class="kpi-sub">\uc704\ud0c1 \uc81c\uc678</span></div>
    <div class="kpi-card green"><span class="kpi-label">\uc720\ub3d9\uc790\uc0b0</span><span class="kpi-value">${fmt(liquidAsset)}</span><span class="kpi-sub">\ube44\uc720\ub3d9/\uc704\ud0c1 \uc81c\uc678</span></div>
    <div class="kpi-card yellow"><span class="kpi-label">\ube44\uc720\ub3d9 \uc790\uc0b0</span><span class="kpi-value">${fmt(illiquidTotal)}</span><span class="kpi-sub">\ubd80\ub3d9\uc0b0 \ub4f1</span></div>
    <div class="kpi-card purple"><span class="kpi-label">\uc704\ud0c1 \uc790\uc0b0</span><span class="kpi-value">${fmt(parentTotal)}</span><span class="kpi-sub">\ubd80\ubaa8\ub2d8</span></div>
    <div class="kpi-card ${savingsRate >= 20 ? 'green' : savingsRate >= 10 ? 'yellow' : 'red'}"><span class="kpi-label">\uc800\ucd95\ub960</span><span class="kpi-value">${savingsRate.toFixed(1)}%</span><span class="kpi-sub">\uc6d4\ud3c9\uade0 \uc218\uc785 ${fmt(avgInc)} / \uc9c0\ucd9c ${fmt(avgExp)}</span></div>
    <div class="kpi-card"><span class="kpi-label">\uc2e0\uc6a9\uc810\uc218</span><span class="kpi-value">${DATA.customer.creditScore}</span><span class="kpi-sub">KCB</span></div>
  `;

  // Monthly chart
  destroyChart('chart-monthly');
  const labels = DATA.months.map(m => m.substring(5));
  charts['chart-monthly'] = new Chart(document.getElementById('chart-monthly'), {
    type: 'bar',
    data: {
      labels,
      datasets: [
        { label: '\uc218\uc785', data: inc, backgroundColor: 'rgba(63,185,80,.7)', borderRadius: 4, order: 2 },
        { label: '\uc9c0\ucd9c', data: exp, backgroundColor: 'rgba(248,81,73,.7)', borderRadius: 4, order: 2 },
        { label: '\uc21c\uc218\uc785', data: inc.map((v, i) => v - exp[i]), type: 'line', borderColor: '#58a6ff', backgroundColor: 'transparent', tension: 0.3, pointRadius: 3, order: 1 }
      ]
    },
    options: { responsive: true, onClick: (e, els) => { if (els.length) { const mi = els[0].index; const di = els[0].datasetIndex; const month = DATA.months.find(m => m.endsWith('-' + labels[mi])) || ''; drillDown({ type: di === 0 ? '\uc218\uc785' : di === 1 ? '\uc9c0\ucd9c' : '', month }); } }, plugins: { legend: { labels: { color: '#8b949e' } } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } }
  });

  // Category pie with period filter
  const filterDiv = document.getElementById('exp-pie-filter');
  if (filterDiv) filterDiv.innerHTML = buildExpPieFilterHtml(expPieFilter);
  renderExpPieOnly();

  // Savings rate trend - bar chart with color-coded bars + data labels
  destroyChart('chart-savings-rate');
  const rates = inc.map((v, i) => v > 100000 ? parseFloat(((v - exp[i]) / v * 100).toFixed(1)) : null);
  const barLabelPlugin = { id:'barLbl', afterDatasetsDraw(chart) { const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,j)=>{ const v=chart.data.datasets[0].data[j]; if(v===null||v===undefined)return; ctx.save();ctx.fillStyle='#e6edf3';ctx.font='bold 11px system-ui';ctx.textAlign='center';ctx.textBaseline='bottom';ctx.fillText(v.toFixed(1)+'%',bar.x,bar.y-3);ctx.restore(); }); } };
  charts['chart-savings-rate'] = new Chart(document.getElementById('chart-savings-rate'), {
    type: 'bar',
    data: { labels, datasets: [
      { label: '\uc800\ucd95\ub960 (%)', data: rates, backgroundColor: rates.map(v => v===null ? 'rgba(139,148,158,.3)' : v>=20 ? 'rgba(63,185,80,.7)' : v>=10 ? 'rgba(210,153,34,.7)' : 'rgba(248,81,73,.7)'), borderRadius: 4 },
      { type: 'line', label: '\ubaa9\ud45c 20%', data: labels.map(() => 20), borderColor: 'rgba(63,185,80,.7)', borderDash: [5,5], borderWidth: 2, pointRadius: 0, tension: 0, fill: false }
    ]},
    options: { responsive: true, plugins: { legend: { labels: { color: '#8b949e', filter: item => item.datasetIndex === 1 } } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => v + '%' }, grid: { color: '#30363d44' } } } },
    plugins: [barLabelPlugin]
  });

  // Latest month summary widget (replaces recent transactions)
  const latestMi = (() => { for (let i = inc.length-1; i >= 0; i--) { if (inc[i] > 50000 || exp[i] > 50000) return i; } return inc.length-1; })();
  const prevMi = latestMi > 0 ? latestMi - 1 : null;
  const mInc = inc[latestMi], mExp = exp[latestMi], mFcf = mInc - mExp;
  const mRate = mInc > 0 ? (mFcf / mInc * 100) : 0;
  const pInc = prevMi !== null ? inc[prevMi] : null, pExp = prevMi !== null ? exp[prevMi] : null;
  const pFcf = prevMi !== null ? pInc - pExp : null, pRate = (prevMi !== null && pInc > 0) ? ((pInc-pExp)/pInc*100) : null;
  const delta = (cur, prev, lowerBetter=false) => {
    if (prev === null || prev === 0) return '<span class="delta neutral">-</span>';
    const pct = (cur - prev) / prev * 100, up = pct > 0;
    const cls = lowerBetter ? (up ? 'up-bad' : 'down-good') : (up ? 'up-good' : 'down-bad');
    return '<span class="delta ' + cls + '">' + (up ? '\u25b2' : '\u25bc') + ' ' + Math.abs(pct).toFixed(1) + '%</span>';
  };
  const cf2 = getEditedCashFlow();
  const topExp = Object.entries(cf2.expense).map(([k,v]) => [k, v[latestMi]||0]).filter(([,v]) => v>0).sort((a,b) => b[1]-a[1]).slice(0,5);
  const maxE = topExp.length ? topExp[0][1] : 1;
  document.getElementById('month-summary-title').textContent = DATA.months[latestMi] + ' \ud604\ud669';
  document.getElementById('month-summary-content').innerHTML =
    '<div class="month-compare-grid">' +
    '<div class="compare-kpi"><span class="lbl">\uc218\uc785</span><span class="val amount-pos">' + fmt(mInc) + '</span>' + delta(mInc, pInc) + '</div>' +
    '<div class="compare-kpi"><span class="lbl">\uc9c0\ucd9c</span><span class="val amount-neg">' + fmt(mExp) + '</span>' + delta(mExp, pExp, true) + '</div>' +
    '<div class="compare-kpi"><span class="lbl">FCF</span><span class="val ' + (mFcf>=0?'amount-pos':'amount-neg') + '">' + fmt(mFcf) + '</span>' + delta(mFcf, pFcf) + '</div>' +
    '<div class="compare-kpi"><span class="lbl">\uc800\ucd95\ub960</span><span class="val ' + (mRate>=20?'amount-pos':mRate>=10?'':'amount-neg') + '">' + mRate.toFixed(1) + '%</span>' + (pRate!==null ? delta(mRate, pRate) : '<span class="delta neutral">-</span>') + '</div>' +
    '</div>' +
    '<div style="font-size:12px;color:var(--text2);margin-bottom:8px">\uc774\ub2ec \uc9c0\ucd9c TOP 5</div>' +
    topExp.map(([cat,v]) => '<div class="bar-h"><span class="bar-label" title="' + cat + '">' + (cat.length>5?cat.substring(0,5)+'\u2026':cat) + '</span><div class="bar-track"><div class="bar-fill" style="width:' + (v/maxE*100).toFixed(1) + '%;background:var(--red)"></div></div><span class="bar-value">' + fmt(v) + '</span></div>').join('') +
    (topExp.length===0 ? '<div style="color:var(--text2);text-align:center;padding:20px">\ub370\uc774\ud130 \uc5c6\uc74c</div>' : '');
}

// ===== Cash Flow =====
function renderCashflow() {
  const cf = getEditedCashFlow();
  const inc = DATA.months.map((_, i) => Object.values(cf.income).reduce((s, v) => s + (v[i] || 0), 0));
  const exp = DATA.months.map((_, i) => Object.values(cf.expense).reduce((s, v) => s + (v[i] || 0), 0));
  const labels = DATA.months.map(m => m.substring(5));
  const totalInc = inc.reduce((a, b) => a + b, 0);
  const totalExp = exp.reduce((a, b) => a + b, 0);
  const avgInc = totalInc / Math.max(inc.filter(v => v > 0).length, 1);
  const avgExp = totalExp / Math.max(exp.filter(v => v > 0).length, 1);

  document.getElementById('kpi-cashflow').innerHTML = `
    <div class="kpi-card green"><span class="kpi-label">\ucd1d \uc218\uc785 (1\ub144)</span><span class="kpi-value">${fmt(totalInc)}</span></div>
    <div class="kpi-card red"><span class="kpi-label">\ucd1d \uc9c0\ucd9c (1\ub144)</span><span class="kpi-value">${fmt(totalExp)}</span></div>
    <div class="kpi-card green"><span class="kpi-label">\uc6d4\ud3c9\uade0 \uc218\uc785</span><span class="kpi-value">${fmt(avgInc)}</span></div>
    <div class="kpi-card red"><span class="kpi-label">\uc6d4\ud3c9\uade0 \uc9c0\ucd9c</span><span class="kpi-value">${fmt(avgExp)}</span></div>
    <div class="kpi-card blue"><span class="kpi-label">\uc5f0\uac04 \uc21c\uc218\uc785</span><span class="kpi-value">${fmt(totalInc - totalExp)}</span></div>
  `;

  // Stacked bar - income and expense as absolute values side by side
  destroyChart('chart-cashflow-detail');
  const incDatasets = Object.entries(cf.income).filter(([, v]) => v.some(x => x > 0)).map(([k, v], i) => ({
    label: k, data: v, backgroundColor: COLORS[i % COLORS.length], stack: 'income'
  }));
  const expDatasets = Object.entries(cf.expense).filter(([, v]) => v.some(x => x > 0)).map(([k, v], i) => ({
    label: k, data: v, backgroundColor: COLORS[(i + 4) % COLORS.length], stack: 'expense'
  }));
  const expRatioPlugin = { id:'expRatio', afterDatasetsDraw(chart) { const ctx=chart.ctx; const meta0=chart.getDatasetMeta(0), meta1=chart.getDatasetMeta(1); if(!meta0||!meta1)return; const n=chart.data.labels.length; for(let i=0;i<n;i++){const bar0=meta0.data[i],bar1=meta1.data[i];if(!bar0||!bar1)continue;const inc=chart.data.datasets[0].data[i],exp=Object.values(cf.expense).reduce((s,v)=>s+(v[i]||0),0);const pct=inc>0?(exp/inc*100):0;const x=(bar0.x+bar1.x)/2,y=Math.min(bar0.y,bar1.y)-8;ctx.save();ctx.fillStyle='#d29922';ctx.font='bold 10px system-ui';ctx.textAlign='center';ctx.fillText(pct.toFixed(0)+'%',x,y);ctx.restore();} } };
  charts['chart-cashflow-detail'] = new Chart(document.getElementById('chart-cashflow-detail'), {
    type: 'bar',
    data: { labels, datasets: [...incDatasets, ...expDatasets] },
    options: { responsive: true, onClick: (e, els) => { if (els.length) { const el = els[0]; const cat = [...incDatasets, ...expDatasets][el.datasetIndex].label; const month = DATA.months.find(m => m.endsWith('-' + labels[el.index])) || ''; const type = el.datasetIndex < incDatasets.length ? '\uc218\uc785' : '\uc9c0\ucd9c'; drillDown({ type, category: cat, month }); } }, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } },
    plugins: [expRatioPlugin]
  });

  // Income breakdown - uses edited data
  destroyChart('chart-income-breakdown');
  const incTotals = Object.entries(cf.income).map(([k, v]) => [k, v.reduce((a, b) => a + b, 0)]).filter(([, v]) => v > 0);
  charts['chart-income-breakdown'] = new Chart(document.getElementById('chart-income-breakdown'), {
    type: 'doughnut',
    data: { labels: incTotals.map(([k]) => k), datasets: [{ data: incTotals.map(([, v]) => v), backgroundColor: COLORS }] },
    options: { responsive: true, onClick: (e, els) => { if (els.length) { drillDown({ type: '\uc218\uc785', category: incTotals[els[0].index][0] }); } }, plugins: { legend: { position: 'bottom', labels: { color: '#8b949e' } } } },
    plugins: [pieLblPlugin]
  });

  // Expense breakdown (monthly average) - uses edited data
  destroyChart('chart-expense-breakdown');
  const avgE = getAvgExpenseByCategory();
  const expSorted = Object.entries(avgE).filter(([, v]) => v > 0).sort((a, b) => b[1] - a[1]).slice(0, 10);
  charts['chart-expense-breakdown'] = new Chart(document.getElementById('chart-expense-breakdown'), {
    type: 'bar',
    data: { labels: expSorted.map(([k]) => k), datasets: [{ data: expSorted.map(([, v]) => Math.round(v)), backgroundColor: COLORS, borderRadius: 6 }] },
    options: { indexAxis: 'y', responsive: true, onClick: (e, els) => { if (els.length) { drillDown({ type: '\uc9c0\ucd9c', category: expSorted[els[0].index][0] }); } }, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } }, y: { ticks: { color: '#8b949e' }, grid: { display: false } } } },
    plugins: [hBarFmtPlugin]
  });

  // Cashflow table - uses edited data
  let html = '<thead><tr><th>\ud56d\ubaa9</th>';
  DATA.months.forEach(m => html += `<th>${m.substring(5)}</th>`);
  html += '<th>\ud569\uacc4</th></tr></thead><tbody>';
  html += '<tr style="background:rgba(63,185,80,.05)"><td><strong>\uc218\uc785</strong></td>';
  inc.forEach(v => html += `<td class="amount-pos">${fmt(v)}</td>`);
  html += `<td class="amount-pos"><strong>${fmt(totalInc)}</strong></td></tr>`;
  for (const [cat, vals] of Object.entries(cf.income)) {
    const sum = vals.reduce((a, b) => a + b, 0);
    if (sum === 0) continue;
    html += `<tr><td style="padding-left:20px">${cat}</td>`;
    vals.forEach(v => html += `<td>${fmt(v)}</td>`);
    html += `<td>${fmt(sum)}</td></tr>`;
  }
  html += '<tr style="background:rgba(248,81,73,.05)"><td><strong>\uc9c0\ucd9c</strong></td>';
  exp.forEach(v => html += `<td class="amount-neg">${fmt(v)}</td>`);
  html += `<td class="amount-neg"><strong>${fmt(totalExp)}</strong></td></tr>`;
  for (const [cat, vals] of Object.entries(cf.expense)) {
    const sum = vals.reduce((a, b) => a + b, 0);
    if (sum === 0) continue;
    html += `<tr><td style="padding-left:20px">${cat}</td>`;
    vals.forEach(v => html += `<td>${fmt(v)}</td>`);
    html += `<td>${fmt(sum)}</td></tr>`;
  }
  html += '<tr style="border-top:2px solid var(--blue)"><td><strong>\uc21c\uc218\uc785</strong></td>';
  inc.forEach((v, i) => { const net = v - exp[i]; html += `<td class="${amtClass(net)}"><strong>${fmt(net)}</strong></td>`; });
  html += `<td class="${amtClass(totalInc - totalExp)}"><strong>${fmt(totalInc - totalExp)}</strong></td></tr>`;
  html += '</tbody>';
  document.getElementById('cashflow-table').innerHTML = html;
}

// ===== Assets =====
function renderAssets() {
  const myTotal = getTotalAssets(true, false);
  const liquidTotal = getTotalAssets(true, true);
  const allTotal = getTotalAssets(false, false);
  const parentTotal = allTotal - myTotal;
  const illiquidTotal = myTotal - liquidTotal;

  document.getElementById('kpi-assets').innerHTML = `
    <div class="kpi-card blue"><span class="kpi-label">\ub0b4 \ucd1d\uc790\uc0b0</span><span class="kpi-value">${fmt(myTotal)}</span><span class="kpi-sub">\uc720\ub3d9 + \ube44\uc720\ub3d9 (\uc704\ud0c1 \uc81c\uc678)</span></div>
    <div class="kpi-card green"><span class="kpi-label">\uc720\ub3d9\uc790\uc0b0</span><span class="kpi-value">${fmt(liquidTotal)}</span><span class="kpi-sub">\uc989\uc2dc \ud604\uae08\ud654 \uac00\ub2a5\ud55c \uc790\uc0b0</span></div>
    <div class="kpi-card yellow"><span class="kpi-label">\ube44\uc720\ub3d9 \uc790\uc0b0</span><span class="kpi-value">${fmt(illiquidTotal)}</span><span class="kpi-sub">\ucd1d\uc790\uc0b0\uc5d0 \ud3ec\ud568, \ud604\uae08\ud654 \uc5b4\ub824\uc6c0</span></div>
    <div class="kpi-card purple"><span class="kpi-label">\uc704\ud0c1 \uc790\uc0b0</span><span class="kpi-value">${fmt(parentTotal)}</span><span class="kpi-sub">\ubd80\ubaa8\ub2d8 \uc790\uc0b0 (\ub0b4 \uc790\uc0b0 \uc544\ub2d8)</span></div>
  `;

  // Group assets
  const groups = {};
  DATA.assets.forEach(a => {
    if (!groups[a.category]) groups[a.category] = [];
    groups[a.category].push(a);
  });

  let html = '';
  for (const [cat, items] of Object.entries(groups)) {
    const catTotal = items.reduce((s, a) => s + (isParent(a.id) ? 0 : a.amount), 0);
    html += `<div class="asset-group"><div class="asset-group-title"><span>${cat}</span><span style="color:var(--blue)">${fmt(catTotal)}</span></div>`;
    items.forEach(a => {
      const pd = isParent(a.id);
      const il = isIlliquid(a.id);
      html += `<div class="asset-item ${pd ? 'parent-delegated' : ''} ${il ? 'illiquid' : ''}">
        <div class="left"><div class="name" title="${a.name}">${a.name}</div></div>
        <div class="amount ${a.amount > 0 ? '' : 'amount-zero'}">${a.amount >= 100000000 ? fmt(a.amount) : fmtFull(a.amount)}</div>
        <div class="asset-toggles">
          <label class="toggle-label parent" title="\uccb4\ud06c\uc2dc \ub0b4 \ucd1d\uc790\uc0b0\uc5d0\uc11c \uc644\uc804\ud788 \uc81c\uc678"><input type="checkbox" ${pd ? 'checked' : ''} onchange="toggleParent('${a.id}')"> \uc704\ud0c1(\uc81c\uc678)</label>
          <label class="toggle-label illiquid" title="\uccb4\ud06c\uc2dc \ucd1d\uc790\uc0b0\uc5d0 \ud3ec\ud568\ub418\uc9c0\ub9cc \uc720\ub3d9\uc790\uc0b0\uc5d0\uc11c \uc81c\uc678"><input type="checkbox" ${il ? 'checked' : ''} onchange="toggleIlliquid('${a.id}')"> \ube44\uc720\ub3d9</label>
        </div>
      </div>`;
    });
    html += '</div>';
  }
  document.getElementById('asset-list').innerHTML = html;

  // Asset allocation chart (my liquid assets only)
  destroyChart('chart-asset-alloc');
  const catTotals = {};
  DATA.assets.forEach(a => {
    if (isParent(a.id) || isIlliquid(a.id)) return;
    catTotals[a.category] = (catTotals[a.category] || 0) + a.amount;
  });
  const allocData = Object.entries(catTotals).filter(([, v]) => v > 0).sort((a, b) => b[1] - a[1]);
  charts['chart-asset-alloc'] = new Chart(document.getElementById('chart-asset-alloc'), {
    type: 'doughnut',
    data: { labels: allocData.map(([k]) => k), datasets: [{ data: allocData.map(([, v]) => Math.round(v)), backgroundColor: COLORS }] },
    options: { responsive: true, plugins: { legend: { position: 'bottom', labels: { color: '#8b949e', font: { size: 11 } } } } },
    plugins: [pieLblPlugin]
  });

  // Investment assets detail
  destroyChart('chart-invest-alloc');
  const invAssets = DATA.assets.filter(a => a.category === '\ud22c\uc790\uc131 \uc790\uc0b0' && a.amount > 0 && !isParent(a.id));
  if (invAssets.length) {
    charts['chart-invest-alloc'] = new Chart(document.getElementById('chart-invest-alloc'), {
      type: 'doughnut',
      data: { labels: invAssets.map(a => a.name.substring(0, 15)), datasets: [{ data: invAssets.map(a => Math.round(a.amount)), backgroundColor: COLORS }] },
      options: { responsive: true, plugins: { legend: { position: 'bottom', labels: { color: '#8b949e', font: { size: 10 } } } } },
      plugins: [pieLblPlugin]
    });
  }
}

// ===== Transactions =====
let txPage = 1;
const TX_PER_PAGE = 30;
let filteredTx = [];

function renderTransactions() {
  const allTx = getAllTx();
  const cats = [...new Set(allTx.map(t => t.category))].sort();
  const catSel = document.getElementById('tx-category');
  const prevCat = catSel.value;
  catSel.innerHTML = '<option value="">\ubaa8\ub4e0 \uce74\ud14c\uace0\ub9ac</option>';
  cats.forEach(c => { if (c) { const o = document.createElement('option'); o.value = c; o.text = c; catSel.add(o); } });
  if (prevCat && [...catSel.options].some(o => o.value === prevCat)) catSel.value = prevCat;
  const months = [...new Set(allTx.map(t => t.date.substring(0, 7)))].sort().reverse();
  const monthSel = document.getElementById('tx-month');
  const prevMonth = monthSel.value;
  monthSel.innerHTML = '<option value="">\ubaa8\ub4e0 \uc6d4</option>';
  months.forEach(m => { const o = document.createElement('option'); o.value = m; o.text = m; monthSel.add(o); });
  if (prevMonth && [...monthSel.options].some(o => o.value === prevMonth)) monthSel.value = prevMonth;
  // Populate category datalist for edit modal
  const dl = document.getElementById('cat-list');
  dl.innerHTML = '';
  cats.forEach(c => { if (c) { const o = document.createElement('option'); o.value = c; dl.appendChild(o); } });
  // Show reset button if edits exist
  document.getElementById('btn-reset-edits').style.display = Object.keys(txEdits).length > 0 ? '' : 'none';
  filterTx();
}

function filterTx() {
  const search = document.getElementById('tx-search').value.toLowerCase();
  const type = document.getElementById('tx-type').value;
  const cat = document.getElementById('tx-category').value;
  const month = document.getElementById('tx-month').value;
  filteredTx = getAllTx().filter(t => {
    if (search && !t.description.toLowerCase().includes(search) && !t.category.toLowerCase().includes(search) && !t.subcategory.toLowerCase().includes(search) && !t.paymentMethod.toLowerCase().includes(search)) return false;
    if (type && t.type !== type) return false;
    if (cat && t.category !== cat) return false;
    if (month && !t.date.startsWith(month)) return false;
    return true;
  });
  txPage = 1;
  renderTxTable();
  renderTxCharts();
}

function renderTxCharts() {
  const month = document.getElementById('tx-month').value;
  const container = document.getElementById('tx-charts-container');
  if (filteredTx.length === 0) {
    container.style.display = 'none';
    destroyChart('chart-tx-expense'); destroyChart('chart-tx-monthly');
    return;
  }
  container.style.display = '';
  const expByCat = {}, txByMonth = {};
  let totalInc = 0, totalExp = 0;
  filteredTx.forEach(t => {
    const m = t.date.substring(0, 7);
    if (!txByMonth[m]) txByMonth[m] = {inc: 0, exp: 0};
    if (t.type === '\uc9c0\ucd9c') { expByCat[t.category] = (expByCat[t.category] || 0) + Math.abs(t.amount); totalExp += Math.abs(t.amount); txByMonth[m].exp += Math.abs(t.amount); }
    else if (t.type === '\uc218\uc785') { totalInc += t.amount; txByMonth[m].inc += t.amount; }
  });
  const expSorted = Object.entries(expByCat).sort((a, b) => b[1] - a[1]).slice(0, 8);
  const monthKeys = Object.keys(txByMonth).sort();
  const suffix = month ? ' \u2013 ' + month : (monthKeys.length === 1 ? ' \u2013 ' + monthKeys[0] : '');
  document.getElementById('tx-chart-exp-title').textContent = '\uce74\ud14c\uace0\ub9ac\ubcc4 \uc9c0\ucd9c' + suffix;
  document.getElementById('tx-chart-mon-title').textContent = '\uc6d4\ubcc4 \uc218\uc785/\uc9c0\ucd9c' + (suffix ? suffix : '');
  destroyChart('chart-tx-expense');
  if (expSorted.length > 0) {
    const pieLbl = { id:'pLbl', afterDatasetsDraw(chart) { const ctx=chart.ctx,ds=chart.data.datasets[0],tot=ds.data.reduce((a,b)=>a+b,0); chart.getDatasetMeta(0).data.forEach((arc,j)=>{ const v=ds.data[j]; if(!v||v/tot<0.05)return; const mid=arc.startAngle+(arc.endAngle-arc.startAngle)/2,r=(arc.innerRadius+arc.outerRadius)/2,x=arc.x+Math.cos(mid)*r,y=arc.y+Math.sin(mid)*r; ctx.save();ctx.fillStyle='#fff';ctx.font='bold 10px system-ui';ctx.textAlign='center';ctx.textBaseline='middle';ctx.fillText(fmt(v),x,y);ctx.restore(); }); } };
    charts['chart-tx-expense'] = new Chart(document.getElementById('chart-tx-expense'), {
      type: 'doughnut',
      data: { labels: expSorted.map(([k]) => k), datasets: [{ data: expSorted.map(([, v]) => Math.round(v)), backgroundColor: COLORS }] },
      options: { responsive: true, onClick: (e, els) => { if (els.length) drillDown({ type: '\uc9c0\ucd9c', category: expSorted[els[0].index][0], month: month || undefined }); }, plugins: { legend: { position: 'right', labels: { color: '#8b949e', padding: 6, font: { size: 10 } } } } },
      plugins: [pieLbl]
    });
  }
  destroyChart('chart-tx-monthly');
  if (monthKeys.length > 0) {
    charts['chart-tx-monthly'] = new Chart(document.getElementById('chart-tx-monthly'), {
      type: 'bar',
      data: { labels: monthKeys.map(m => m.substring(5)), datasets: [
        { label: '\uc218\uc785', data: monthKeys.map(m => txByMonth[m].inc), backgroundColor: 'rgba(63,185,80,.7)', borderRadius: 4 },
        { label: '\uc9c0\ucd9c', data: monthKeys.map(m => txByMonth[m].exp), backgroundColor: 'rgba(248,81,73,.7)', borderRadius: 4 }
      ]},
      options: { responsive: true, plugins: { legend: { labels: { color: '#8b949e', font: { size: 11 } } } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } }
    });
  }
}

function renderTxTable() {
  const start = (txPage - 1) * TX_PER_PAGE;
  const page = filteredTx.slice(start, start + TX_PER_PAGE);
  const totalPages = Math.ceil(filteredTx.length / TX_PER_PAGE);

  document.getElementById('tx-count').textContent = `${filteredTx.length}\uac74`;
  let html = '<thead><tr><th>\ub0a0\uc9dc</th><th>\uc2dc\uac04</th><th>\ud0c0\uc785</th><th>\ub300\ubd84\ub958</th><th>\uc18c\ubd84\ub958</th><th>\ub0b4\uc6a9</th><th>\uae08\uc561</th><th></th></tr></thead><tbody>';
  page.forEach(t => {
    const edited = t._edited ? '<span class="badge badge-edited">\uc218\uc815\ub428</span>' : '';
    html += `<tr style="${t._edited ? 'background:rgba(210,153,34,.06)' : ''}" ondblclick="openEditModal(${t._idx})"><td>${t.date}</td><td>${t.time.substring(0, 5)}</td><td><span class="badge ${badgeClass(t.type)}">${t.type}</span>${edited}</td><td>${t.category}</td><td>${t.subcategory}</td><td>${t.description}</td><td class="${amtClass(t.amount)}">${fmtFull(t.amount)}</td><td><button class="btn btn-sm" onclick="openEditModal(${t._idx})">&#9998;</button></td></tr>`;
  });
  html += '</tbody>';
  document.getElementById('tx-table').innerHTML = html;

  let pHtml = `<button ${txPage <= 1 ? 'disabled' : ''} onclick="txPage=1;renderTxTable()">&laquo;</button>`;
  pHtml += `<button ${txPage <= 1 ? 'disabled' : ''} onclick="txPage--;renderTxTable()">&lsaquo;</button>`;
  const startP = Math.max(1, txPage - 2), endP = Math.min(totalPages, txPage + 2);
  for (let i = startP; i <= endP; i++) {
    pHtml += `<button class="${i === txPage ? 'active' : ''}" onclick="txPage=${i};renderTxTable()">${i}</button>`;
  }
  pHtml += `<button ${txPage >= totalPages ? 'disabled' : ''} onclick="txPage++;renderTxTable()">&rsaquo;</button>`;
  pHtml += `<button ${txPage >= totalPages ? 'disabled' : ''} onclick="txPage=${totalPages};renderTxTable()">&raquo;</button>`;
  pHtml += `<span class="info">${txPage}/${totalPages}</span>`;
  document.getElementById('tx-pagination').innerHTML = pHtml;
}

// Edit modal
function openEditModal(idx) {
  const t = getTx(idx);
  document.getElementById('edit-idx').value = idx;
  document.getElementById('edit-date').value = t.date;
  document.getElementById('edit-type').value = t.type;
  document.getElementById('edit-category').value = t.category;
  document.getElementById('edit-subcategory').value = t.subcategory;
  document.getElementById('edit-description').value = t.description;
  document.getElementById('edit-amount').value = t.amount;
  document.getElementById('btn-revert').style.display = txEdits[idx] ? '' : 'none';
  document.getElementById('edit-overlay').classList.add('show');
}
function closeEditModal() { document.getElementById('edit-overlay').classList.remove('show'); }
function saveEdit() {
  const idx = parseInt(document.getElementById('edit-idx').value);
  const orig = DATA.transactions[idx];
  const edits = {};
  const newDate = document.getElementById('edit-date').value;
  const newType = document.getElementById('edit-type').value;
  const newCat = document.getElementById('edit-category').value;
  const newSub = document.getElementById('edit-subcategory').value;
  const newDesc = document.getElementById('edit-description').value;
  const newAmt = parseFloat(document.getElementById('edit-amount').value) || 0;
  if (newDate !== orig.date) edits.date = newDate;
  if (newType !== orig.type) edits.type = newType;
  if (newCat !== orig.category) edits.category = newCat;
  if (newSub !== orig.subcategory) edits.subcategory = newSub;
  if (newDesc !== orig.description) edits.description = newDesc;
  if (newAmt !== orig.amount) edits.amount = newAmt;
  if (Object.keys(edits).length > 0) {
    txEdits[idx] = edits;
    localStorage.setItem(lsKey('txEdits'), JSON.stringify(txEdits));
  }
  closeEditModal();
  renderTransactions();
}
function revertEdit() {
  const idx = parseInt(document.getElementById('edit-idx').value);
  delete txEdits[idx];
  localStorage.setItem(lsKey('txEdits'), JSON.stringify(txEdits));
  closeEditModal();
  renderTransactions();
}
function resetAllEdits() {
  showConfirm('\ubaa8\ub4e0 \uc218\uc815 \ub0b4\uc6a9\uc744 \ucd08\uae30\ud654\ud558\uc2dc\uaca0\uc2b5\ub2c8\uae4c?\n\uc774 \uc791\uc5c5\uc740 \ub418\ub3cc\ub9b4 \uc218 \uc5c6\uc2b5\ub2c8\ub2e4.', () => {
    txEdits = {};
    localStorage.setItem(lsKey('txEdits'), JSON.stringify(txEdits));
    renderTransactions();
  });
}
function exportTxCsv() {
  const headers = ['\ub0a0\uc9dc', '\uc2dc\uac04', '\ud0c0\uc785', '\ub300\ubd84\ub958', '\uc18c\ubd84\ub958', '\ub0b4\uc6a9', '\uae08\uc561', '\uacb0\uc81c\uc218\ub2e8', '\uba54\ubaa8'];
  const rows = [headers];
  filteredTx.forEach(t => rows.push([t.date, t.time ? t.time.substring(0,5) : '', t.type, t.category, t.subcategory, t.description, t.amount, t.paymentMethod || '', t.memo || '']));
  const csv = rows.map(r => r.map(f => '"' + String(f===null||f===undefined?'':f).replace(/"/g,'""') + '"').join(',')).join('\r\n');
  const blob = new Blob(['\ufeff' + csv], {type:'text/csv;charset=utf-8'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'transactions_' + new Date().toISOString().substring(0,10) + '.csv';
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

document.getElementById('tx-search').addEventListener('input', filterTx);
document.getElementById('tx-type').addEventListener('change', filterTx);
document.getElementById('tx-category').addEventListener('change', filterTx);
document.getElementById('tx-month').addEventListener('change', filterTx);

// ===== Investments =====
function renderInvestments() {
  const allInv = DATA.investments;
  // Build type filter
  const types = [...new Set(allInv.map(i => i.type || '\uae30\ud0c0'))].filter(t => t).sort();
  const typeList = ['\uc804\uccb4', ...types];
  document.getElementById('inv-type-filter').innerHTML = '<span class="lbl">\ubd84\ub958:</span>' +
    typeList.map(t => '<button class="type-btn' + ((t==='\uc804\uccb4'&&!currentInvType)||(t===currentInvType)?'  active':'') + '" onclick="currentInvType=\'' + (t==='\uc804\uccb4'?'':t) + '\';renderInvestments()">' + t + '</button>').join('');

  // Filter with original indices preserved
  const invWithIdx = allInv.map((i, idx) => ({...i, _idx: idx}));
  const filtered = currentInvType ? invWithIdx.filter(i => (i.type || '\uae30\ud0c0') === currentInvType) : invWithIdx;
  const activeFiltered = filtered.filter(i => !isExcludedInv(i._idx));
  const excludedFiltered = filtered.filter(i => isExcludedInv(i._idx));

  const totalPrincipal = activeFiltered.reduce((s, i) => s + i.principal, 0);
  const totalCurrent = activeFiltered.reduce((s, i) => s + i.currentValue, 0);
  const totalReturn = totalPrincipal > 0 ? ((totalCurrent - totalPrincipal) / totalPrincipal * 100) : 0;
  const pnl = totalCurrent - totalPrincipal;
  const exPrincipal = excludedFiltered.reduce((s, i) => s + i.principal, 0);
  const exCurrent = excludedFiltered.reduce((s, i) => s + i.currentValue, 0);
  const typeLabel = currentInvType ? ' (' + currentInvType + ')' : '';

  document.getElementById('kpi-investments').innerHTML =
    '<div class="kpi-card blue"><span class="kpi-label">\ud22c\uc790\uc6d0\uae08' + typeLabel + '</span><span class="kpi-value">' + fmt(totalPrincipal) + '</span></div>' +
    '<div class="kpi-card ' + (totalCurrent >= totalPrincipal ? 'green' : 'red') + '"><span class="kpi-label">\ud3c9\uac00\uae08\uc561' + typeLabel + '</span><span class="kpi-value">' + fmt(totalCurrent) + '</span></div>' +
    '<div class="kpi-card ' + (pnl >= 0 ? 'green' : 'red') + '"><span class="kpi-label">\ucd1d \uc190\uc775</span><span class="kpi-value">' + fmtPct(totalReturn) + '</span><span class="kpi-sub">' + fmt(pnl) + '</span></div>' +
    (excludedFiltered.length ? '<div class="kpi-card purple"><span class="kpi-label">\uc81c\uc678\ub41c \ud22c\uc790</span><span class="kpi-value">' + fmt(exCurrent) + '</span><span class="kpi-sub">\uc6d0\uae08 ' + fmt(exPrincipal) + '</span></div>' : '');

  // Portfolio chart (active filtered)
  destroyChart('chart-portfolio');
  const withValue = activeFiltered.filter(i => i.currentValue > 0);
  charts['chart-portfolio'] = new Chart(document.getElementById('chart-portfolio'), {
    type: 'doughnut',
    data: { labels: withValue.map(i => i.name.substring(0, 20)), datasets: [{ data: withValue.map(i => Math.round(i.currentValue)), backgroundColor: COLORS }] },
    options: { responsive: true, plugins: { legend: { position: 'bottom', labels: { color: '#8b949e', font: { size: 10 } } } } },
    plugins: [pieLblPlugin]
  });

  // Returns chart (filtered)
  destroyChart('chart-returns');
  charts['chart-returns'] = new Chart(document.getElementById('chart-returns'), {
    type: 'bar',
    data: {
      labels: filtered.map(i => i.name.substring(0, 12)),
      datasets: [{ data: filtered.map(i => i.returnRate), backgroundColor: filtered.map(i => isExcludedInv(i._idx) ? 'rgba(139,148,158,.4)' : i.returnRate >= 0 ? 'rgba(63,185,80,.7)' : 'rgba(248,81,73,.7)'), borderRadius: 4 }]
    },
    options: { indexAxis: 'y', responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e', callback: v => v + '%' }, grid: { color: '#30363d44' } }, y: { ticks: { color: '#8b949e', font: { size: 10 } }, grid: { display: false } } } },
    plugins: [pctBarPlugin]
  });

  // Investment list (filtered)
  let html = '';
  filtered.forEach(i => {
    const iPnl = i.currentValue - i.principal;
    const ex = isExcludedInv(i._idx);
    html += '<div class="inv-card" style="' + (ex ? 'opacity:.5' : '') + '">' +
      '<div style="display:flex;justify-content:space-between;align-items:start">' +
      '<div><div class="inv-name">' + i.name + '</div><div class="inv-company">' + i.company + ' | ' + i.type + '</div></div>' +
      '<label class="toggle-label"><input type="checkbox" ' + (ex ? 'checked' : '') + ' onchange="toggleExcludedInv(' + i._idx + ')"> \uc81c\uc678</label>' +
      '</div><div class="inv-stats">' +
      '<div class="inv-stat"><span class="label">\ud22c\uc790\uc6d0\uae08</span><span class="value">' + fmtFull(i.principal) + '</span></div>' +
      '<div class="inv-stat"><span class="label">\ud3c9\uac00\uae08\uc561</span><span class="value">' + fmtFull(i.currentValue) + '</span></div>' +
      '<div class="inv-stat"><span class="label">\uc218\uc775\ub960</span><span class="value ' + (i.returnRate >= 0 ? 'amount-pos' : 'amount-neg') + '">' + fmtPct(i.returnRate) + '</span></div>' +
      '<div class="inv-stat"><span class="label">\uc190\uc775\uae08</span><span class="value ' + (iPnl >= 0 ? 'amount-pos' : 'amount-neg') + '">' + fmtFull(iPnl) + '</span></div>' +
      (i.startDate ? '<div class="inv-stat"><span class="label">\uac00\uc785\uc77c</span><span class="value">' + i.startDate + '</span></div>' : '') +
      '</div></div>';
  });
  if (filtered.length === 0) html = '<div style="color:var(--text2);text-align:center;padding:40px">' + (currentInvType||'\uc120\ud0dd\ub41c') + ' \uc720\ud615\uc758 \ud22c\uc790 \uc5c6\uc74c</div>';
  document.getElementById('investment-list').innerHTML = html;
}

// ===== Insurance =====
function renderInsurance() {
  const ins = DATA.insurance;
  const totalPaid = ins.reduce((s, i) => s + i.totalPaid, 0);

  document.getElementById('kpi-insurance').innerHTML = `
    <div class="kpi-card blue"><span class="kpi-label">\ubcf4\uc720 \ubcf4\ud5d8</span><span class="kpi-value">${ins.length}\uac74</span></div>
    <div class="kpi-card"><span class="kpi-label">\ucd1d \ub0a9\uc785\uae08</span><span class="kpi-value">${fmt(totalPaid)}</span></div>
  `;

  let html = '';
  ins.forEach(i => {
    html += `<div class="ins-card">
      <div class="ins-left"><div class="ins-name">${i.name}</div><div class="ins-company">${i.company} | ${i.startDate} ~ ${i.endDate}</div></div>
      <div class="ins-right"><span class="ins-status">${i.status}</span><div class="ins-paid">\ub0a9\uc785: ${fmtFull(i.totalPaid)}</div></div>
    </div>`;
  });
  document.getElementById('insurance-list').innerHTML = html;
}

// ===== Analysis =====
function renderAnalysis() {
  const inc = getMonthlyIncome();
  const exp = getMonthlyExpense();
  const avgInc = inc.filter(v => v > 0).reduce((a, b) => a + b, 0) / Math.max(inc.filter(v => v > 0).length, 1);
  const avgExp = exp.filter(v => v > 0).reduce((a, b) => a + b, 0) / Math.max(exp.filter(v => v > 0).length, 1);
  const savingsRate = avgInc > 0 ? ((avgInc - avgExp) / avgInc * 100) : 0;
  const myAssets = getTotalAssets(true);
  const emergencyMonths = avgExp > 0 ? (DATA.assets.filter(a => !isParent(a.id) && (a.category === '\uc790\uc720\uc785\ucd9c\uae08 \uc790\uc0b0' || a.category === '\uc800\ucd95\uc131 \uc790\uc0b0')).reduce((s, a) => s + a.amount, 0) / avgExp) : 0;

  // Health score (0-100)
  let score = 0;
  score += Math.min(savingsRate / 30 * 30, 30); // savings rate: max 30
  score += Math.min(emergencyMonths / 6 * 25, 25); // emergency fund: max 25
  score += DATA.customer.creditScore >= 900 ? 20 : DATA.customer.creditScore >= 800 ? 15 : 10; // credit: max 20
  const invDiv = DATA.investments.filter(i => i.currentValue > 0).length;
  score += Math.min(invDiv / 5 * 15, 15); // diversification: max 15
  score += DATA.insurance.length >= 2 ? 10 : 5; // insurance: max 10
  score = Math.round(score);

  const scoreClass = score >= 75 ? 'good' : score >= 50 ? 'ok' : 'bad';
  const scoreLabel = score >= 75 ? '\uc6b0\uc218' : score >= 50 ? '\ubcf4\ud1b5' : '\uac1c\uc120 \ud544\uc694';

  // Top 5 expense categories
  const avgE = getAvgExpenseByCategory();
  const top5 = Object.entries(avgE).filter(([, v]) => v > 0).sort((a, b) => b[1] - a[1]).slice(0, 5);
  const maxVal = top5.length ? top5[0][1] : 1;

  let barsHtml = '';
  top5.forEach(([cat, val], i) => {
    barsHtml += `<div class="bar-h"><span class="bar-label">${cat}</span><div class="bar-track"><div class="bar-fill" style="width:${(val/maxVal*100).toFixed(1)}%;background:${COLORS[i]}"></div></div><span class="bar-value">${fmt(val)}</span></div>`;
  });

  // Fixed vs variable expenses
  const fixedCats = ['\uad50\uc721/\ud559\uc2b5', '\uae08\uc735', '\uc8fc\uac70/\ud1b5\uc2e0', '\uc790\ub140/\uc721\uc544'];
  const fixedTotal = fixedCats.reduce((s, c) => s + (avgE[c] || 0), 0);
  const variableTotal = avgExp - fixedTotal;

  document.getElementById('analysis-content').innerHTML = `
    <div class="analysis-card">
      <h3>\uc7ac\ubb34 \uac74\uac15 \uc810\uc218</h3>
      <div class="health-score ${scoreClass}">${score}<span style="font-size:20px">/100</span></div>
      <div style="text-align:center;color:var(--text2);margin-bottom:12px">${scoreLabel}</div>
      <div style="font-size:12px;color:var(--text2);line-height:1.8">
        <div>\uc800\ucd95\ub960: ${savingsRate.toFixed(1)}% ${savingsRate >= 20 ? '\u2705' : '\u26a0\ufe0f'}</div>
        <div>\ube44\uc0c1\uae08: ${emergencyMonths.toFixed(1)}\uac1c\uc6d4\ubd84 ${emergencyMonths >= 6 ? '\u2705' : '\u26a0\ufe0f'}</div>
        <div>\uc2e0\uc6a9\uc810\uc218: ${DATA.customer.creditScore} ${DATA.customer.creditScore >= 900 ? '\u2705' : '\u26a0\ufe0f'}</div>
        <div>\ud22c\uc790 \ub2e4\uac01\ud654: ${invDiv}\uac1c \uc0c1\ud488 ${invDiv >= 5 ? '\u2705' : '\u26a0\ufe0f'}</div>
        <div>\ubcf4\ud5d8: ${DATA.insurance.length}\uac74 \u2705</div>
      </div>
    </div>
    <div class="analysis-card">
      <h3>\uc6d4\ud3c9\uade0 TOP 5 \uc9c0\ucd9c \uce74\ud14c\uace0\ub9ac</h3>
      ${barsHtml}
    </div>
    <div class="analysis-card">
      <h3>\uace0\uc815\uc9c0\ucd9c vs \ubcc0\ub3d9\uc9c0\ucd9c</h3>
      <div style="display:flex;justify-content:center;gap:30px;margin:20px 0">
        <div style="text-align:center"><div style="font-size:12px;color:var(--text2)">\uace0\uc815\uc9c0\ucd9c</div><div style="font-size:22px;font-weight:700;color:var(--orange)">${fmt(fixedTotal)}</div><div style="font-size:11px;color:var(--text2)">${(fixedTotal/avgExp*100).toFixed(0)}%</div></div>
        <div style="text-align:center"><div style="font-size:12px;color:var(--text2)">\ubcc0\ub3d9\uc9c0\ucd9c</div><div style="font-size:22px;font-weight:700;color:var(--blue)">${fmt(variableTotal)}</div><div style="font-size:11px;color:var(--text2)">${(variableTotal/avgExp*100).toFixed(0)}%</div></div>
      </div>
      <div style="font-size:12px;color:var(--text2)">\uace0\uc815: \uad50\uc721/\ud559\uc2b5, \uae08\uc735, \uc8fc\uac70/\ud1b5\uc2e0, \uc790\ub140/\uc721\uc544</div>
    </div>
  `;

  // Category trend chart (top 5 categories monthly)
  destroyChart('chart-category-trend');
  const labels = DATA.months.map(m => m.substring(5));
  const cfData = getEditedCashFlow();
  const trendDatasets = top5.map(([cat], i) => ({
    label: cat,
    data: cfData.expense[cat] || [],
    borderColor: COLORS[i],
    backgroundColor: 'transparent',
    tension: 0.3,
    pointRadius: 2,
  }));
  charts['chart-category-trend'] = new Chart(document.getElementById('chart-category-trend'), {
    type: 'line',
    data: { labels, datasets: trendDatasets },
    options: { responsive: true, plugins: { legend: { labels: { color: '#8b949e' } } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } }
  });

  // Day of week pattern
  destroyChart('chart-day-pattern');
  const dayTotals = [0, 0, 0, 0, 0, 0, 0];
  const dayCounts = [0, 0, 0, 0, 0, 0, 0];
  const dayLabels = ['\uc77c', '\uc6d4', '\ud654', '\uc218', '\ubaa9', '\uae08', '\ud1a0'];
  getAllTx().filter(t => t.type === '\uc9c0\ucd9c' && t.amount < 0).forEach(t => {
    const d = new Date(t.date).getDay();
    dayTotals[d] += Math.abs(t.amount);
    dayCounts[d]++;
  });
  const dayAvg = dayTotals.map((t, i) => dayCounts[i] ? Math.round(t / dayCounts[i]) : 0);
  charts['chart-day-pattern'] = new Chart(document.getElementById('chart-day-pattern'), {
    type: 'bar',
    data: { labels: dayLabels, datasets: [{ label: '\ud3c9\uade0 \uc9c0\ucd9c', data: dayAvg, backgroundColor: dayLabels.map((_, i) => i === 0 || i === 6 ? 'rgba(248,81,73,.7)' : 'rgba(88,166,255,.7)'), borderRadius: 6 }] },
    options: { responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { display: false } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } },
    plugins: [barFmtPlugin]
  });

  // Anomaly detection
  const anomalies = [];
  for (const [cat, vals] of Object.entries(cfData.expense)) {
    const nonZero = vals.filter(v => v > 0);
    if (nonZero.length < 2) continue;
    const avg = nonZero.reduce((a, b) => a + b, 0) / nonZero.length;
    vals.forEach((v, i) => {
      if (v > avg * 2 && v > 50000) {
        anomalies.push({ month: DATA.months[i], category: cat, amount: v, average: avg, ratio: (v / avg).toFixed(1) });
      }
    });
  }
  anomalies.sort((a, b) => b.amount - a.amount);

  let anomalyHtml = '<thead><tr><th>\uc6d4</th><th>\uce74\ud14c\uace0\ub9ac</th><th>\uae08\uc561</th><th>\uc6d4\ud3c9\uade0</th><th>\ubc30\uc728</th></tr></thead><tbody>';
  if (anomalies.length === 0) {
    anomalyHtml += '<tr><td colspan="5" style="text-align:center;color:var(--text2)">\uc774\uc0c1 \uc9c0\ucd9c\uc774 \uac10\uc9c0\ub418\uc9c0 \uc54a\uc558\uc2b5\ub2c8\ub2e4.</td></tr>';
  } else {
    anomalies.slice(0, 15).forEach(a => {
      anomalyHtml += `<tr><td>${a.month}</td><td>${a.category}</td><td class="amount-neg">${fmtFull(a.amount)}</td><td>${fmtFull(a.average)}</td><td style="color:var(--yellow)">${a.ratio}x</td></tr>`;
    });
  }
  anomalyHtml += '</tbody>';
  document.getElementById('anomaly-table').innerHTML = anomalyHtml;

  // Build filter UI
  const filterDiv = document.getElementById('analysis-chart-filter');
  if (filterDiv) filterDiv.innerHTML = buildAnalysisChartFilterHtml(analysisChartFilter);

  renderAnalysisChartsOnly();
}
function renderAnalysisChartsOnly() {
  const {indices, startDate, endDate} = getAnalysisChartTxRange();
  const suffix = analysisChartFilter==='' ? '' : analysisChartFilter==='1' ? ' (최근 1개월)' : analysisChartFilter==='3' ? ' (최근 3개월)' : analysisChartFilter==='6' ? ' (상반기)' : ' (' + analysisChartFilter + ')';

  // Payment method breakdown - filtered by date range
  destroyChart('chart-payment-method');
  const pmData = {};
  getAllTx().filter(t => t.type === '\uc9c0\ucd9c' && (!startDate || (t.date >= startDate && t.date <= endDate))).forEach(t => {
    const pm = t.paymentMethod || '\uae30\ud0c0';
    pmData[pm] = (pmData[pm] || 0) + Math.abs(t.amount);
  });
  const pmSorted = Object.entries(pmData).filter(([,v]) => v > 0).sort((a, b) => b[1] - a[1]).slice(0, 10);
  if (pmSorted.length > 0) {
    const pmTitle = document.querySelector('#page-analysis .chart-grid .chart-card:first-child h3');
    if (pmTitle) pmTitle.textContent = '\uacb0\uc81c\uc218\ub2e8\ubcc4 \uc9c0\ucd9c' + suffix;
    charts['chart-payment-method'] = new Chart(document.getElementById('chart-payment-method'), {
      type: 'bar',
      data: { labels: pmSorted.map(([k]) => k), datasets: [{ data: pmSorted.map(([, v]) => Math.round(v)), backgroundColor: COLORS, borderRadius: 6 }] },
      options: { indexAxis: 'y', responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } }, y: { ticks: { color: '#8b949e', font: { size: 11 } }, grid: { display: false } } } },
      plugins: [hBarFmtPlugin]
    });
  }

  // Hourly spending pattern - filtered by date range
  destroyChart('chart-hour-pattern');
  const hourTotals = new Array(24).fill(0);
  const hourCounts = new Array(24).fill(0);
  getAllTx().filter(t => t.type === '\uc9c0\ucd9c' && t.time && t.time.length >= 2 && (!startDate || (t.date >= startDate && t.date <= endDate))).forEach(t => {
    const h = parseInt(t.time.substring(0, 2));
    if (!isNaN(h) && h >= 0 && h < 24) { hourTotals[h] += Math.abs(t.amount); hourCounts[h]++; }
  });
  if (hourCounts.some(c => c > 0)) {
    const hrTitle = document.querySelector('#page-analysis .chart-grid .chart-card:last-child h3');
    if (hrTitle) hrTitle.textContent = '\uc2dc\uac04\ub300\ubcc4 \uc9c0\ucd9c \ud328\ud134' + suffix;
    charts['chart-hour-pattern'] = new Chart(document.getElementById('chart-hour-pattern'), {
      type: 'bar',
      data: { labels: Array.from({length: 24}, (_, i) => i + '\uc2dc'), datasets: [{ data: hourTotals.map((t, i) => hourCounts[i] ? Math.round(t / hourCounts[i]) : 0), backgroundColor: Array.from({length: 24}, (_, i) => i >= 6 && i < 22 ? 'rgba(88,166,255,.6)' : 'rgba(188,140,255,.5)'), borderRadius: 4 }] },
      options: { responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e', font: { size: 9 } }, grid: { display: false } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } },
      plugins: [barFmtPlugin]
    });
  }
}

// ===== FCF =====
function renderFcf() {
  const inc = getMonthlyIncome();
  const exp = getMonthlyExpense();
  const fcf = inc.map((v, i) => v - exp[i]);
  const cumFcf = []; fcf.reduce((acc, v) => { cumFcf.push(acc + v); return acc + v; }, 0);
  const totalFcf = fcf.reduce((a, b) => a + b, 0);
  const activeMonths = fcf.filter((_, i) => inc[i] > 100000 || exp[i] > 100000);
  const avgFcf = activeMonths.length ? totalFcf / activeMonths.length : 0;
  const posMonths = fcf.filter(v => v > 0).length;
  const negMonths = fcf.filter(v => v < 0).length;
  const labels = DATA.months.map(m => m.substring(5));

  document.getElementById('kpi-fcf').innerHTML = `
    <div class="kpi-card ${totalFcf >= 0 ? 'green' : 'red'}"><span class="kpi-label">\uc5f0\uac04 FCF</span><span class="kpi-value">${fmt(totalFcf)}</span></div>
    <div class="kpi-card ${avgFcf >= 0 ? 'green' : 'red'}"><span class="kpi-label">\uc6d4\ud3c9\uade0 FCF</span><span class="kpi-value">${fmt(avgFcf)}</span></div>
    <div class="kpi-card green"><span class="kpi-label">\ud751\uc790 \uc6d4</span><span class="kpi-value">${posMonths}\uac1c\uc6d4</span></div>
    <div class="kpi-card red"><span class="kpi-label">\uc801\uc790 \uc6d4</span><span class="kpi-value">${negMonths}\uac1c\uc6d4</span></div>
  `;

  destroyChart('chart-fcf-monthly');
  charts['chart-fcf-monthly'] = new Chart(document.getElementById('chart-fcf-monthly'), {
    type: 'bar',
    data: { labels, datasets: [{ label: 'FCF', data: fcf, backgroundColor: fcf.map(v => v >= 0 ? 'rgba(63,185,80,.7)' : 'rgba(248,81,73,.7)'), borderRadius: 4 }] },
    options: { responsive: true, onClick: (e, els) => { if (els.length) { const month = DATA.months.find(m => m.endsWith('-' + labels[els[0].index])) || ''; drillDown({ month }); } }, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } },
    plugins: [barFmtPlugin]
  });

  destroyChart('chart-fcf-cumulative');
  charts['chart-fcf-cumulative'] = new Chart(document.getElementById('chart-fcf-cumulative'), {
    type: 'line',
    data: { labels, datasets: [{ label: '\ub204\uc801 FCF', data: cumFcf, borderColor: '#58a6ff', backgroundColor: 'rgba(88,166,255,.1)', fill: true, tension: 0.3, pointRadius: 4 }] },
    options: { responsive: true, plugins: { legend: { display: false } }, scales: { x: { ticks: { color: '#8b949e' }, grid: { color: '#30363d22' } }, y: { ticks: { color: '#8b949e', callback: v => fmt(v) }, grid: { color: '#30363d44' } } } },
    plugins: [lineLblPlugin]
  });

  document.getElementById('fcf-advice').innerHTML = generateFcfAdvice(inc, exp, fcf, avgFcf);
}

function generateFcfAdvice(inc, exp, fcf, avgFcf) {
  const avgInc = inc.filter(v => v > 100000).reduce((a, b) => a + b, 0) / Math.max(inc.filter(v => v > 100000).length, 1);
  const avgExp = exp.filter(v => v > 100000).reduce((a, b) => a + b, 0) / Math.max(exp.filter(v => v > 100000).length, 1);
  const liquidAsset = getTotalAssets(true, true);
  const emergencyMonths = avgExp > 0 ? liquidAsset / avgExp : 0;
  const savingsRate = avgInc > 0 ? ((avgInc - avgExp) / avgInc * 100) : 0;
  let html = '';

  // Card 1: Asset Allocation Guide
  html += '<div class="analysis-card"><h3>\ud83d\udcb0 \uc790\uc0b0\ubc30\ubd84 \uac00\uc774\ub4dc</h3><div style="font-size:13px;line-height:2.2">';
  html += `<div>\uc6d4\ud3c9\uade0 FCF: <strong>${fmt(avgFcf)}</strong></div>`;
  html += `<div>\ube44\uc0c1\uae08: <strong>${emergencyMonths.toFixed(1)}\uac1c\uc6d4\ubd84</strong> (\ud604\uc7ac \uc720\ub3d9\uc790\uc0b0 ${fmt(liquidAsset)})</div>`;
  if (emergencyMonths < 3) {
    html += '<div style="color:var(--red)">\ud83d\udea8 \ube44\uc0c1\uae08 \ucd5c\uc6b0\uc120 \ud655\ubcf4 \ud544\uc694 (\ucd5c\uc18c 3~6\uac1c\uc6d4\ubd84)</div>';
    html += `<div>\uad8c\uc7a5 \ubc30\ubd84: FCF\uc758 <strong>80%</strong> \u2192 \ube44\uc0c1\uae08, <strong>20%</strong> \u2192 \uc800\ucd95</div>`;
    html += `<div>\ubaa9\ud45c \ube44\uc0c1\uae08: ${fmt(avgExp * 6)} (6\uac1c\uc6d4\ubd84)</div>`;
  } else if (emergencyMonths < 6) {
    html += '<div style="color:var(--yellow)">\u26a0\ufe0f \ube44\uc0c1\uae08 \ubcf4\ucda9 \ud544\uc694 (\ubaa9\ud45c: 6\uac1c\uc6d4\ubd84)</div>';
    html += '<div>\uad8c\uc7a5 \ubc30\ubd84: FCF\uc758 <strong>50%</strong> \u2192 \ube44\uc0c1\uae08, <strong>30%</strong> \u2192 \ud22c\uc790, <strong>20%</strong> \u2192 \uc800\ucd95</div>';
  } else {
    html += '<div style="color:var(--green)">\u2705 \ube44\uc0c1\uae08 \ucda9\ubd84 (\ubaa9\ud45c \ub2ec\uc131)</div>';
    html += '<div>\uad8c\uc7a5 \ubc30\ubd84: FCF\uc758 <strong>60%</strong> \u2192 \ud22c\uc790, <strong>20%</strong> \u2192 \ucd94\uac00\uc800\ucd95, <strong>20%</strong> \u2192 \uc790\uae30\uacc4\ubc1c</div>';
  }
  if (avgFcf > 0) {
    html += '<div style="margin-top:8px;padding-top:8px;border-top:1px solid var(--border)">';
    html += `<div>\ud22c\uc790 \uad8c\uc7a5\uc561 (\uc6d4): <strong style="color:var(--blue)">${fmt(Math.round(avgFcf * (emergencyMonths >= 6 ? 0.6 : emergencyMonths >= 3 ? 0.3 : 0)))}</strong></div>`;
    html += `<div>\uc800\ucd95 \uad8c\uc7a5\uc561 (\uc6d4): <strong style="color:var(--green)">${fmt(Math.round(avgFcf * (emergencyMonths >= 6 ? 0.2 : emergencyMonths >= 3 ? 0.2 : 0.2)))}</strong></div>`;
    html += '</div>';
  }
  html += '</div></div>';

  // Card 2: Product Recommendations
  html += '<div class="analysis-card"><h3>\ud83d\udcca \ucd94\ucc9c \uae08\uc735\uc0c1\ud488</h3><div style="font-size:13px;line-height:2.2">';
  if (avgFcf <= 0) {
    html += '<div style="color:var(--text2)">\uc6d4\ud3c9\uade0 FCF\uac00 \uc801\uc790\uc785\ub2c8\ub2e4. \uc9c0\ucd9c \uc808\uac10\uc774 \uc6b0\uc120\uc785\ub2c8\ub2e4.</div>';
  } else {
    if (emergencyMonths < 6) {
      html += '<div><strong style="color:var(--green)">\ube44\uc0c1\uae08 \uc6b0\uc120</strong></div>';
      html += '<div>\u2022 \ud30c\ud0b9\ud1b5\uc7a5: \ud1a0\uc2a4\ubc45\ud06c/\uce74\uce74\uc624\ubc45\ud06c (\uc5f0 2%\ub300, \uc989\uc2dc \ucd9c\uae08)</div>';
      html += '<div>\u2022 CMA: \uc99d\uad8c\uc0ac CMA (\uc5f0 3%\ub300, \uc790\uc720\uc785\ucd9c\uae08)</div>';
    }
    if (avgFcf >= 500000) {
      html += '<div style="margin-top:6px"><strong style="color:var(--blue)">\ud22c\uc790 \uc0c1\ud488 (\uc6d4 50\ub9cc\uc6d0+ FCF)</strong></div>';
      html += '<div>\u2022 <strong>KODEX 200</strong>: \uad6d\ub0b4 \ub300\ud615\uc8fc ETF (\uc7a5\uae30 \ud22c\uc790)</div>';
      html += '<div>\u2022 <strong>TIGER \ubbf8\uad6dS&P500</strong>: \ubbf8\uad6d \ubd84\uc0b0\ud22c\uc790 ETF</div>';
      html += '<div>\u2022 <strong>KODEX \ubbf8\uad6d\ub098\uc2a4\ub2e5100</strong>: \ubbf8\uad6d \uae30\uc220\uc8fc ETF</div>';
      html += '<div>\u2022 <strong>KOSEF \uad6d\uace0\ucc2810\ub144</strong>: \ucc44\uad8c\ud615 ETF (\uc548\uc815)</div>';
    }
    if (avgFcf >= 300000) {
      html += '<div style="margin-top:6px"><strong style="color:var(--yellow)">\uc815\uae30\uc800\ucd95</strong></div>';
      html += `<div>\u2022 \uc801\uae08 \uc790\ub3d9\uc774\uccb4: \uc6d4 ${fmt(Math.round(avgFcf * 0.2))} \uad8c\uc7a5</div>`;
    }
    if (avgFcf >= 1000000) {
      html += '<div style="margin-top:6px"><strong style="color:var(--purple)">\uace0\uc561 \ud22c\uc790\uc790 (\uc6d4 100\ub9cc+ FCF)</strong></div>';
      html += '<div>\u2022 \uac1c\ubcc4\uc8fc\uc2dd \ud3ec\ud2b8\ud3f4\ub9ac\uc624 \uad6c\uc131 \uace0\ub824</div>';
      html += '<div>\u2022 ISA \uacc4\uc88c \ud65c\uc6a9 (\uc808\uc138 \ud6a8\uacfc)</div>';
      html += '<div>\u2022 \uc5f0\uae08\uc800\ucd95\ud380\ub4dc/IRP (\uc5f0\ub9d0\uc815\uc0b0 \uc138\uc561\uacf5\uc81c)</div>';
    }
  }
  html += '</div></div>';

  // Card 3: Warnings
  html += '<div class="analysis-card"><h3>\u26a0\ufe0f \uc8fc\uc758\uc0ac\ud56d</h3><div style="font-size:13px;line-height:2.2">';
  const warnings = [];
  if (emergencyMonths < 3) warnings.push('\ud83d\udea8 \ube44\uc0c1\uae08 \ubd80\uc871: \ucd5c\uc18c 3\uac1c\uc6d4 \uc0dd\ud65c\ube44 \ud655\ubcf4 \uc2dc\uae09');
  if (savingsRate < 10) warnings.push('\ud83d\udea8 \uc800\ucd95\ub960 \ub9e4\uc6b0 \ub0ae\uc74c (' + savingsRate.toFixed(1) + '%). \ubaa9\ud45c: 20%');
  else if (savingsRate < 20) warnings.push('\u26a0\ufe0f \uc800\ucd95\ub960 \uac1c\uc120 \ud544\uc694 (' + savingsRate.toFixed(1) + '%). \ubaa9\ud45c: 20%');
  const negCount = fcf.filter(v => v < 0).length;
  if (negCount >= 3) warnings.push('\u26a0\ufe0f \uc801\uc790\uc6d4 ' + negCount + '\uac1c\uc6d4: \uc9c0\ucd9c \uad6c\uc870 \uc810\uac80 \ud544\uc694');
  const expVals = exp.filter(v => v > 100000);
  if (expVals.length >= 2) {
    const expStd = Math.sqrt(expVals.reduce((s, v) => s + (v - avgExp) ** 2, 0) / (expVals.length - 1));
    if (avgExp > 0 && expStd / avgExp > 0.3) warnings.push('\u26a0\ufe0f \uc6d4\ubcc4 \uc9c0\ucd9c \ud3b8\ucc28 \ud07c (\ubcc0\ub3d9\uacc4\uc218 ' + (expStd / avgExp * 100).toFixed(0) + '%). \uc608\uc0b0 \uad00\ub9ac \ud544\uc694');
  }
  const avgE2 = getAvgExpenseByCategory();
  const topCat = Object.entries(avgE2).sort((a, b) => b[1] - a[1])[0];
  if (topCat && avgExp > 0 && topCat[1] / avgExp > 0.3) warnings.push('\u26a0\ufe0f \uc9c0\ucd9c \uc9d1\uc911: ' + topCat[0] + '\uc774 \uc6d4\ud3c9\uade0 \uc9c0\ucd9c\uc758 ' + (topCat[1] / avgExp * 100).toFixed(0) + '% \ucc28\uc9c0');
  if (warnings.length === 0) warnings.push('\u2705 \ud604\uc7ac \uc7ac\ubb34 \uc0c1\ud0dc \uc591\ud638\ud569\ub2c8\ub2e4.');
  warnings.forEach(w => html += `<div>${w}</div>`);
  html += '</div></div>';

  return html;
}

// ===== Init =====
document.getElementById('user-name').textContent = DATA.customer.name;
document.getElementById('user-detail').textContent = DATA.customer.gender + ' / ' + DATA.customer.age + '\uc138';
document.getElementById('gen-info').textContent = '\uc0dd\uc131: ' + DATA.generatedAt;
updatePersonToggle();

// File upload events
const dropZone = document.getElementById('drop-zone');
const fileInput = document.getElementById('file-input');
dropZone.addEventListener('click', () => fileInput.click());
dropZone.addEventListener('dragover', (e) => { e.preventDefault(); dropZone.classList.add('dragover'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop', (e) => { e.preventDefault(); dropZone.classList.remove('dragover'); handleFileUpload(e.dataTransfer.files); });
fileInput.addEventListener('change', (e) => { handleFileUpload(e.target.files); e.target.value = ''; });

renderDashboard();
</script>
</body>
</html>'''


def main():
    paths = sys.argv[1:] if len(sys.argv) > 1 else [DEFAULT_EXCEL]
    all_data = []
    for i, path in enumerate(paths):
        if not os.path.exists(path):
            print(f'Error: File not found: {path}')
            sys.exit(1)
        label = f'Person {i + 1}' if len(paths) > 1 else ''
        print(f'Reading{" " + label if label else ""}: {path}')
        data = extract_data(path)
        print(f'  - {len(data["transactions"])} transactions')
        print(f'  - {len(data["assets"])} assets')
        print(f'  - {len(data["months"])} months of cash flow')
        all_data.append(data)

    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard.html')
    html = generate_html(all_data)
    with open(output, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'Dashboard generated: {output}')


if __name__ == '__main__':
    main()
